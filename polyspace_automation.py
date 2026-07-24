#coding=utf8
import os
import time
import re
import subprocess
import json
import multiprocessing
import configparser
import sys
import shutil
from colorama import Fore, init
from subprocess import Popen,PIPE,STDOUT
from openpyxl import load_workbook
from lxml import html, etree
from lxml import etree as ET
import datetime
import getpass

sys.path.insert(0,os.getcwd())

polyspaceWorkMode = 'normal'
polyspaceAnalysisMode = 'specify'
polyspacechangeFilePath = ''

# compiler select, e.g. cmake, emake
compilerEnv = 'cmake'

init(autoreset = True)

class Util:
    def __init__(self):
        pass

    @staticmethod
    def normalize_path(path):
        """Normalize path and remove duplicate directory names"""
        import re
        if not path:
            return path

        # Convert to normpath first
        normalized = os.path.normpath(path)

        # Remove duplicate directory patterns (e.g., Build/Build, Customer/Customer)
        # This regex matches directory names that are repeated consecutively
        while True:
            # Find patterns like /Build/Build/ or \Build\Build\
            pattern = r'([^\\/]+)[\\/]+\1(?=[\\/]|$)'
            new_path = re.sub(pattern, r'\1', normalized, flags=re.IGNORECASE)
            if new_path == normalized:
                break
            normalized = new_path

        return normalized
    @staticmethod
    def getProjectPath():
        """Get project root directory (compatible with two structures)"""
        current_dir = sys.path[0]

        # Normalize path separators
        current_dir = current_dir.replace('\\', '/')

        # Case 1: Customer/Build/tools/Components/Polyspace structure
        customer_build_pattern = '/Customer/Build/tools/Components/Polyspace'
        if customer_build_pattern in current_dir:
            # Extract everything before /Customer/Build
            parts = current_dir.split(customer_build_pattern)
            return parts[0]

        # Case 2: Customer/Build structure (shorter path)
        if '/Customer/Build/' in current_dir:
            parts = current_dir.split('/Customer/Build/')
            return parts[0]

        # Case 3: Build/tools/Components/Polyspace structure
        build_pattern = '/Build/tools/Components/Polyspace'
        if build_pattern in current_dir:
            parts = current_dir.split(build_pattern)
            return parts[0]

        # Case 4: Build structure (shorter path)
        if '/Build/' in current_dir:
            parts = current_dir.split('/Build/')
            return parts[0]

        # Case 5: Try to find by walking up
        path = current_dir
        while path:
            # Check if this looks like a project root
            if os.path.exists(os.path.join(path, "BSW")) or \
                    os.path.exists(os.path.join(path, "SourceCode")) or \
                    os.path.exists(os.path.join(path, ".git")):
                return path
            parent = os.path.dirname(path)
            if parent == path:  # Reached root
                break
            path = parent

        # Fallback: remove everything after the last Build
        parts = current_dir.split('/')
        for i in range(len(parts) - 1, -1, -1):
            if parts[i] == "Build":
                return '/'.join(parts[:i])

        return os.path.dirname(current_dir)

    @staticmethod
    def getBuildPath():
        """Get Build directory path without duplication"""
        current_dir = sys.path[0]
        current_dir = os.path.normpath(current_dir.replace('\\', '/'))

        # Define possible Build directory patterns
        build_patterns = [
            ('/Customer/Build/', '/Customer/Build'),
            ('/Build/', '/Build'),
            ('\\Customer\\Build\\', '\\Customer\\Build'),
            ('\\Build\\', '\\Build')
        ]

        # Check each pattern
        for pattern, replacement in build_patterns:
            if pattern in current_dir:
                # Extract path before the pattern
                parts = current_dir.split(pattern)
                if parts and parts[0]:
                    base_path = parts[0]
                    # Reconstruct Build path
                    build_path = base_path + replacement
                    build_path = os.path.normpath(build_path)

                    # Verify the path exists
                    if os.path.exists(build_path):
                        return build_path

        # If not found in path patterns, check common locations
        project_root = Util.getProjectPath()

        # Define possible Build locations in priority order
        possible_build_paths = [
            os.path.join(project_root, "Customer", "Build"),
            os.path.join(project_root, "Build"),
            os.path.join(os.path.dirname(project_root), "Build") if project_root else None,
        ]

        for build_path in possible_build_paths:
            if build_path and os.path.exists(build_path):
                return os.path.normpath(build_path)

        # Fallback
        return os.path.join(project_root, "Build")

    @staticmethod
    def getConfigPath():
        """Get configuration file path (compatible with two structures)"""
        build_path = Util.getBuildPath()

        # Try multiple possible locations
        possible_paths = [
            os.path.join(build_path, "VerifyCfg", "config.ini"),
            os.path.join(Util.getProjectPath(), "Build", "VerifyCfg", "config.ini"),
            os.path.join(Util.getProjectPath(), "Customer", "Build", "VerifyCfg", "config.ini"),
        ]

        for config_path in possible_paths:
            if os.path.exists(config_path):
                return config_path

        # If not found, use original path
        return os.path.join(Util.getProjectPath(), 'Build', 'VerifyCfg', 'config.ini')

    @staticmethod
    def getCmakePath():
        """Get Cmake directory path (compatible with two structures)"""
        build_path = Util.getBuildPath()
        return os.path.join(build_path, 'Cmake', 'include.cmake')

    @staticmethod
    def getSourceRelPath():
        if (compilerEnv == 'cmake'):
            relPath = '/SourceCode/'
        elif (compilerEnv == 'emake'):
            relPath = '/'
        else:
            relPath = '/SourceCode/'
        return relPath

    @staticmethod
    def getSourcePath():
        """Get source code base path (compatible with two structures)"""
        project_root = Util.getProjectPath()

        if compilerEnv == 'cmake':
            # Try multiple possible source code directories
            possible_paths = [
                os.path.join(project_root, 'SourceCode'),
                os.path.join(project_root, 'BSW'),
                os.path.join(project_root, 'Customer', 'SourceCode'),
                os.path.join(project_root, 'APP'),
                os.path.join(project_root, 'ASW'),
                project_root,  # Directly use project root directory
            ]

            for path in possible_paths:
                if os.path.exists(path):
                    return path

            # If none found, use original logic
            return os.path.join(project_root, 'SourceCode')
        elif compilerEnv == 'emake':
            return project_root

    @staticmethod
    def reportError(inputStr):
        print(Fore.RED + inputStr)
        exit(-1)

    @staticmethod
    def reportNormal(inputStr):
        print(Fore.GREEN + inputStr)

    @staticmethod
    def resolveRelativePath(rel_path):
        """Resolve relative path to absolute path (intelligently match project structure)"""
        if not rel_path:
            return rel_path

        rel_path = rel_path.strip()

        # If already absolute path, return directly
        if os.path.isabs(rel_path):
            return os.path.normpath(rel_path)

        # Remove "./" or ".\" prefix
        if rel_path.startswith('./') or rel_path.startswith('.\\'):
            rel_path = rel_path[2:]

        # Get project root directory
        project_root = Util.getProjectPath()

        # Case 1: Direct project root directory
        direct_path = os.path.join(project_root, rel_path)

        if os.path.exists(direct_path):
            return os.path.normpath(direct_path)

        # Case 2: Check common subdirectories
        common_subdirs = ['BSW', 'ASW', 'APP', 'SourceCode', 'Customer/SourceCode']

        for subdir in common_subdirs:
            # Handle nested paths
            subdir_path = os.path.join(project_root, *subdir.split('/'))
            test_path = os.path.join(subdir_path, rel_path)

            if os.path.exists(test_path):
                return os.path.normpath(test_path)

        # Case 3: Try to extract directory information from relative path
        # If rel_path is already part of full path, try to match
        path_parts = rel_path.replace('\\', '/').split('/')
        if len(path_parts) >= 2:
            # Try to find matching parent directory
            for root, dirs, files in os.walk(project_root):
                for dir_name in dirs:
                    if dir_name == path_parts[0]:
                        # Build complete path
                        test_path = os.path.join(root, *path_parts)
                        if os.path.exists(test_path):
                            return os.path.normpath(test_path)

        # Case 4: If file is included in projectSource, use path from build_src.txt
        # This logic should be handled in the calling function

        # Finally return the expected path
        return os.path.normpath(direct_path)


class PolyspaceUtil():

    def __init__(self):
        pass

    def start(self):
        # Get the configuration file
        print(self.getPath())
        path = Util.getConfigPath()

        if(not os.path.exists(path)):
            Util.reportError(f'{path} Not Found!!!!')
            raise ValueError

        print(path+" Found")
        self.initConfig(path)

        self.setAllConfig()

        self.clean(0)

        ret2 = self.traversalFile()

        # generate PSPRJ project file
        self.generatePSPRJFile()

        #generate option-command.file
        self.getOptionsCommand()
        print("Launching...")

        self.runLaunchScript()

        #remove exit launching script
        if(os.path.exists(self.polyspacePath+'\\Component\\generate_launching_script.bat')):
            os.remove(self.polyspacePath+'\\Component\\generate_launching_script.bat')
        shutil.move(self.polyspacePath+'\\generate_launching_script.bat', self.polyspacePath+'\\Component')

        self.changePolyspaceProject()
        ret1 = self.analysisReport(ret2)
        self.generateJsonFile()

    def generateJsonFile(self):
        print(subprocess.run(f'{self.matlabPath}/polyspace/bin/polyspace-results-export.exe -format csv -results-dir {self.polyspacePath}/Component/CP_Result -output-name  {self.polyspacePath}/Component/CP_Result/Polyspace-Doc/Polyspace.csv -set-language-english'))

    def clean(self, flag):
        print(f'polyspace path: {self.polyspacePath}')
        if(os.path.exists(self.getProject()) and flag == 0):
            os.remove(self.getProject())
            print(Fore.GREEN + "Clean "+ self.getProject()+ " Success")

        if(os.path.exists(self.polyspacePath+'/Component') and flag == 0):
            shutil.rmtree(self.polyspacePath+'/Component')
            print(Fore.GREEN + "Clean "+ self.componentPath + " Success")

        if(os.path.exists(self.polyspacePath+'\\Polyspace')):
            shutil.rmtree(self.polyspacePath+'\\Polyspace')
            print(Fore.GREEN + "Clean "+ self.polyspacePath+'\\Polyspace' + " Success")

        if(os.path.exists(self.polyspacePath+'\\generate_launching_script.bat') and flag == 0):
            os.remove(self.polyspacePath+'\\generate_launching_script.bat')
            print(Fore.GREEN + "Clean "+ self.polyspacePath+'\\generate_launching_script.bat  Success')


        if(os.path.exists(self.polyspacePath+'/Component/Project.json') and flag == 0):
            os.remove(self.polyspacePath+'/Component/Project.json')
            print(Fore.GREEN + "Clean "+ self.polyspacePath+'/Component/Project.json')

    def changePolyspaceProject(self):
        fd = open(self.polyspacePath + '\\Polyspace.psprj', 'r+', encoding='utf-8')
        text = fd.readlines()
        fd.close()
        fd = open(self.polyspacePath + '\\Polyspace.psprj', 'w+', encoding='utf-8')
        for line in text:
            if(re.compile('.*?(<result>).*?').findall(line) != []):
                fd.write('<result>\n<file path="file:/'+self.launchResult.replace('\\','/')+'/" isactive="true"/> \n</result>\n')
            elif(re.compile('.*?(</result>).*?').findall(line) != []):
                continue
            else:
                fd.write(line)
        fd.close()

    def fileMove(self, dest, src, destDir):
        if(os.path.exists(dest)):
            os.remove(dest)

        shutil.move(src, destDir)

    def processFile(self, fd, filePath):

        for line in fd.readlines():
            ret = re.compile('polyspace<.*?>').findall(line)
            if(ret != []):
                print(ret)
                self.problemSet.add(filePath)
                self.allNumber += 1
        return 0

    def openFile(self, filePath):
        # 直接使用resolveRelativePath
        resolved_path = Util.resolveRelativePath(filePath.replace("\n", ''))
        if (not os.path.exists(resolved_path)):
            print(Fore.RED + resolved_path + " don't exist")
            exit(-1)
        self.allFile.add(resolved_path)

        # 保持原始的文件编码尝试逻辑
        fileDes = None
        try:
            fileDes = open(resolved_path, 'r', encoding="gb2312")
            self.processFile(fileDes, resolved_path)
            fileDes.close()
        except:
            try:
                if fileDes:
                    fileDes.close()
                fileDes = open(resolved_path, 'r', encoding="utf-8")
                self.processFile(fileDes, resolved_path)
                fileDes.close()
            except:
                try:
                    if fileDes:
                        fileDes.close()
                    fileDes = open(resolved_path, 'r', encoding="ANSI")
                    self.processFile(fileDes, resolved_path)
                    fileDes.close()
                except Exception as e:
                    print(Fore.RED + "error: ", e)
                    return

    # Read path information from a configuration file
    def traversalFile(self):
        print("Detect polyspace annotation information for all files")
        fd = open(self.polyspaceCheck, 'r', encoding="utf-8")
        lines = fd.readlines()
        lineKey = False
        for line in lines:
            line = line.replace("\\", "/")
            if line.startswith('[SpecifyAnalysis]') and analysisModeParm == 'specify':
                lineKey = True
                continue
            elif line.startswith('[') and lineKey == True:
                break

            if lineKey:
                if not line.startswith('./'):
                    continue
                # 使用新的路径解析
                resolved_path = Util.resolveRelativePath(line.replace("\n", '').replace('./', ''))
                if (not os.path.exists(resolved_path)):
                    self.noSuchFile.add(resolved_path)
        if (len(self.noSuchFile) != 0):
            for file in self.noSuchFile:
                print(Fore.RED + "No Such File : " + file)
            exit(-1)

        if (len(self.problemSet) == 0):
            return 0
        else:
            return -1

    def getPolyspaceRules(self):
        wb = load_workbook(self.polyspaceRules)

        # 尝试找到合适的sheet
        target_sheet_name = None
        for sheet_name in wb.sheetnames:
            if '报备项' in sheet_name or 'WhiteList' in sheet_name or '规则' in sheet_name:
                target_sheet_name = sheet_name
                break

        if not target_sheet_name:
            # 如果没有找到特定名称的sheet，使用第一个sheet
            target_sheet_name = wb.sheetnames[0]

        # 获取工作表
        ws = wb[target_sheet_name]

        # 读取表头
        firstRow = next(ws.iter_rows(min_row=1, max_row=1))
        headers = [cell.value for cell in firstRow]

        # 查找列索引
        ruleIndex = -1
        moduleIndex = -1
        for index, header in enumerate(headers):
            if header:
                header_str = str(header)
                if 'Polyspace Rule' in header_str or '规则' in header_str:
                    ruleIndex = index
                elif 'WhiteList Module' in header_str or '模块' in header_str:
                    moduleIndex = index

        # 如果没有找到特定列名，假设第一列是规则，第二列是模块
        if ruleIndex == -1 and len(headers) > 0:
            ruleIndex = 0
        if moduleIndex == -1 and len(headers) > 1:
            moduleIndex = 1

        ruleList = []

        if ruleIndex >= 0 and moduleIndex >= 0:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if ruleIndex < len(row) and row[ruleIndex]:
                    rawRules = row[ruleIndex]
                    rawModules = row[moduleIndex] if moduleIndex < len(row) else ''

                    # 确保规则和模块都是字符串
                    if not isinstance(rawRules, str):
                        rawRules = str(rawRules) if rawRules else ''
                    if not isinstance(rawModules, str):
                        rawModules = str(rawModules) if rawModules else ''

                    ruleList.append((rawRules.strip(), rawModules.strip()))

        # 关闭Excel文件
        wb.close()

        return ruleList

    def processErrorsReport(self):
        filePath = self.reportPath + '/Polyspace_Developer.html'

        fd = open(filePath, 'r', encoding='utf-8')
        lines = fd.read()
        fd.close()

        ruleList = self.getPolyspaceRules()

        # Use HTML parser
        tree = etree.HTML(lines)
        reportedCount = 0
        unreportedOrangeCount = 0
        suppressPath = []

        # Add Suppress analysis file
        lineKey = False
        with open(self.polyspaceCheck, 'r', encoding='utf-8')  as f:
            for line in f.readlines():
                line = line.replace("\\", "/")
                if not lineKey:
                    if line.startswith('[SuppressAnalysis]'):
                        lineKey = True
                        continue
                elif lineKey == True and line.startswith(r'['):
                    break
                else:
                    if not line.startswith('./'):
                        continue
                    suppressPath.append(line[2:])

        # Get module whitelist
        moduleWhiteList = self.getModuleWhiteList()

        # Store file-level statistics
        fileStats = {}  # key: 文件名, value: {'orange_count': 总数, 'reported': 已报备数, 'unreported': 未报备数}

        # Locate target h2 tag - Chapter 3 "Unproven Runtime Errors"
        h2Elements = tree.xpath(
            f"//h2[@class='rgSect1Title'][.//span[@class='rgSect1TitleText'][text()='未证明的运行时错误']]"
        )

        for h2Element in h2Elements:
            # Collect subsequent nodes until encountering h1/h2
            nodes = []
            currentNode = h2Element
            for sibling in currentNode.itersiblings():
                if sibling.tag in ('h1', 'h2'):
                    break
                nodes.append(sibling)

            # Traverse nodes to process tables
            i = 0
            while i < len(nodes):
                node = nodes[i]

                # Check if it is tag with class rgTableTitle
                if node.tag == 'p':
                    classAttr = node.get('class')
                    if classAttr and 'rgTableTitle' in classAttr.split():
                        suppressEnable = False
                        filePath = ''
                        try:
                            filePath = node.xpath(".//span[@class='rgTableTitleText']/text()")[0].replace('\\', '/')
                            fileName = filePath.split('/')[-1].strip()

                            # Base filename (remove .c/.h suffix)
                            baseFileName = fileName
                            if baseFileName.endswith('.c'):
                                baseFileName = baseFileName[:-2]
                            elif baseFileName.endswith('.h'):
                                baseFileName = baseFileName[:-2]

                            if baseFileName not in fileStats:
                                fileStats[baseFileName] = {
                                    'orange_count': 0,
                                    'reported': 0,
                                    'unreported': 0,
                                    'file_path': filePath
                                }

                            for suppress in suppressPath:
                                suppress = suppress.rstrip('\n')
                                if suppress.lower() in filePath.rstrip('\n').lower():
                                    suppressEnable = True
                                    break
                        except IndexError:
                            i += 1
                            continue

                        # Check if the next node is an rgUnruledTable table
                        if i + 1 < len(nodes):
                            nextNode = nodes[i + 1]
                            tableClass = nextNode.get('class')
                            if nextNode.tag == 'table' and tableClass and 'rgUnruledTable' in tableClass.split():
                                table = nextNode
                                headers = table.xpath(".//thead/tr[1]/td//text()")
                                try:
                                    commentIndex = headers.index("注释") + 1
                                except ValueError:
                                    commentIndex = -1
                                try:
                                    checkIndex = headers.index("检查") + 1
                                except ValueError:
                                    checkIndex = -1
                                try:
                                    lineIndex = headers.index("行") + 1
                                except ValueError:
                                    lineIndex = -1
                                try:
                                    rowIndex = headers.index("列") + 1
                                except ValueError:
                                    rowIndex = -1

                                # Traverse table rows
                                rows = table.xpath('.//tbody/tr')
                                for idx, row in enumerate(rows):
                                    check = row.xpath(f".//td[{checkIndex}]/p/span/text()")
                                    if check:
                                        check_type = check[0].strip()
                                        comment = '未报备'
                                        isWhitelisted = False

                                        # Increment orange count for this file
                                        fileStats[baseFileName]['orange_count'] += 1

                                        if suppressEnable:
                                            isWhitelisted = True
                                            reportedCount += 1
                                            fileStats[baseFileName]['reported'] += 1
                                            comment = '已报备，白名单类型：第三方代码'
                                        else:
                                            # ORIGINAL COMPLEX RULE MATCHING LOGIC - ADDED BACK
                                            found_in_rule = False
                                            for rawRules, modules in ruleList:
                                                if rawRules:
                                                    # Rules and modules are separated by separators
                                                    for splitGeneralResult in re.split(r'[\\、]', str(rawRules.strip())):
                                                        ruleName = splitGeneralResult
                                                        if '&' in rawRules:
                                                            # Special treatment for "rule1&rule2"
                                                            splitSpecialResult = re.split(r'[&]', str(rawRules.strip()))
                                                            findRuleStatus = False
                                                            for index1, item1 in enumerate(splitSpecialResult):
                                                                if index1 == 0 and item1 == check_type:
                                                                    findRuleStatus = True
                                                                else:
                                                                    if (item1 == check_type) and (lineIndex != -1) and (
                                                                            rowIndex != -1):
                                                                        baseLine = row.xpath(
                                                                            f".//td[{lineIndex}]/p/span/text()")
                                                                        baseRow = row.xpath(
                                                                            f".//td[{rowIndex}]/p/span/text()")

                                                                        for index2, item2 in enumerate(
                                                                                splitSpecialResult):
                                                                            if index1 != index2:
                                                                                if idx >= 1:
                                                                                    preRow = rows[idx - 1]
                                                                                    preCheck = preRow.xpath(
                                                                                        f".//td[{checkIndex}]/p/span/text()")
                                                                                    preLine = preRow.xpath(
                                                                                        f".//td[{lineIndex}]/p/span/text()")
                                                                                    preRow = preRow.xpath(
                                                                                        f".//td[{rowIndex}]/p/span/text()")

                                                                                    if item2 == preCheck[
                                                                                        0].strip() and baseLine == preLine and baseRow == preRow:
                                                                                        findRuleStatus = True
                                                                                        break

                                                                                if (idx + 1) < len(rows):
                                                                                    nextRow = rows[idx + 1]
                                                                                    nextCheck = nextRow.xpath(
                                                                                        f".//td[{checkIndex}]/p/span/text()")
                                                                                    nextLine = nextRow.xpath(
                                                                                        f".//td[{lineIndex}]/p/span/text()")
                                                                                    nextRow = nextRow.xpath(
                                                                                        f".//td[{rowIndex}]/p/span/text()")

                                                                                    if item2 == nextCheck[
                                                                                        0].strip() and baseLine == nextLine and baseRow == nextRow:
                                                                                        findRuleStatus = True
                                                                                        break

                                                                if findRuleStatus:
                                                                    ruleName = item1
                                                                    break

                                                        if ruleName == check_type:
                                                            if modules == 'All':
                                                                found_in_rule = True
                                                            else:
                                                                if modules:
                                                                    for module in re.split(r'[\\、]',
                                                                                           str(modules.strip())):
                                                                        modulePattern = module.strip().replace('*',
                                                                                                               '.*')
                                                                        if re.match(modulePattern,
                                                                                    filePath.rstrip('\n'),
                                                                                    re.IGNORECASE):
                                                                            found_in_rule = True
                                                                            break
                                                            if found_in_rule:
                                                                break

                                                if found_in_rule:
                                                    break

                                            if found_in_rule:
                                                isWhitelisted = True
                                                reportedCount += 1
                                                fileStats[baseFileName]['reported'] += 1
                                                comment = '已报备,白名单类型：' + check_type
                                            else:
                                                # Check if .c file is in module whitelist
                                                if fileName.lower().endswith('.c') and self.isFileInModuleWhiteList(
                                                        fileName, moduleWhiteList):
                                                    isWhitelisted = True
                                                    reportedCount += 1
                                                    fileStats[baseFileName]['reported'] += 1
                                                    comment = '已报备，白名单类型：文件在模块白名单中'

                                        # If not whitelisted, increment unreported orange count
                                        if not isWhitelisted:
                                            unreportedOrangeCount += 1
                                            fileStats[baseFileName]['unreported'] += 1

                                        # Update comment cell
                                        commentCells = row.xpath(f".//td[{commentIndex}]/p")
                                        if commentCells:
                                            commentCell = commentCells[0]
                                            for child in list(commentCell):
                                                commentCell.remove(child)
                                            span = etree.SubElement(commentCell, 'span')
                                            span.text = comment
                                            span.set('style', f'color: {"green" if isWhitelisted else "red"};')

                                i += 2
                                continue
                i += 1

        # Return modified HTML content and statistics
        modifiedHtml = etree.tostring(tree, method='html', encoding='unicode', with_tail=True)
        # Return 3 values, the 3rd is a dictionary containing fileStats
        return (modifiedHtml, reportedCount, {'unreportedCount': unreportedOrangeCount, 'fileStats': fileStats})

    def analysisReport(self, ret2):
        """Analyze and generate the final Polyspace report"""
        # First run report generation
        self.runReportScript()

        # Get red and orange check counts
        red = 0
        orange = 0
        redCount = 0
        orangeCount = 0

        # Process error report to get statistics
        modifiedHtml, reportedCount, statsDict = self.processErrorsReport()

        # Extract statistics
        totalIssues = statsDict['unreportedCount']  # Unreported orange count
        fileStats = statsDict['fileStats']  # File-level statistics
        totalReported = reportedCount  # Reported orange count

        # Re-parse HTML to get runtime check summary
        tree = etree.HTML(modifiedHtml)

        # Locate runtime check summary table (Table 1.3)
        tableTitles = tree.xpath(
            "//p[@class='rgTableTitle'][.//span[@class='rgTableTitleText'][contains(., '{}')]]".format('运行时检查摘要')
        )

        for title in tableTitles:
            nextNode = title.getnext()
            if nextNode is not None and nextNode.tag == 'table' and 'rgUnruledTable' in nextNode.get('class', ''):
                table = nextNode

                # Extract red check count
                red = table.xpath(".//td[p/span[text()='红色检查的数量']]/following-sibling::td[1]/p/span/text()")
                if red:
                    try:
                        redCount = int(red[0].strip())
                    except ValueError:
                        redCount = red[0]

                # Extract orange check count
                orange = table.xpath(".//td[p/span[text()='橙色检查的数量']]/following-sibling::td[1]/p/span/text()")
                if orange:
                    try:
                        orangeCount = int(orange[0].strip())
                    except ValueError:
                        orangeCount = orange[0]

                # Check if red count is 0 and unreported orange is 0
                if (ret2 == 0) and (redCount == 0) and (totalIssues == 0):
                    resultText = f"通过, 红色的数量:{str(redCount)}, 已报备橙色的数量:{str(totalReported)}, 未报备橙色的数量:{str(totalIssues)}"
                    resultColor = "green"
                else:
                    resultText = f"失败, 红色的数量:{str(redCount)}, 已报备橙色的数量:{str(totalReported)}, 未报备橙色的数量:{str(totalIssues)}"
                    resultColor = "red"

                # Update pass/fail cell
                resultCells = table.xpath(
                    ".//td[p/span[text()='通过/失败']]/following-sibling::td[1]/p"
                )
                if resultCells:
                    cell = resultCells[0]
                    # Clear existing content
                    for child in list(cell):
                        cell.remove(child)
                    span = etree.SubElement(cell, 'span')
                    span.text = resultText
                    span.set('style', f'color: {resultColor};')

                # REMOVE THE TWO EXTRA ROWS - DO NOT add "已报备橙色检查的数量" and "未报备橙色检查的数量" rows
                # This section is now commented out to remove those rows

                # Find the orange check row and remove any following rows we might have added before
                rows = table.xpath(".//tbody/tr")
                for i, row in enumerate(rows):
                    cells = row.xpath(".//td")
                    if len(cells) >= 2:
                        firstCell = cells[0]
                        text = firstCell.xpath(".//span/text()")
                        if text and text[0].strip() == "橙色检查的数量":
                            # Check if next rows are the extra rows we want to remove
                            if i + 1 < len(rows):
                                nextRowText = rows[i + 1].xpath(".//td[1]//span/text()")
                                if nextRowText and ("已报备橙色检查的数量" in nextRowText[0] or
                                                    "未报备橙色检查的数量" in nextRowText[0]):
                                    # Remove the next two rows (both extra rows)
                                    parent = rows[i + 1].getparent()
                                    if parent is not None:
                                        # Remove first extra row
                                        parent.remove(rows[i + 1])
                                        if i + 1 < len(parent):  # Check if second row still exists
                                            secondExtraRow = parent[i + 1]
                                            if "未报备橙色检查的数量" in secondExtraRow.xpath(".//td[1]//span/text()")[0]:
                                                parent.remove(secondExtraRow)
                            break

        finalHtml = etree.tostring(tree, method='html', encoding='unicode', with_tail=True)

        filePath = self.reportPath + '/CATL_Polyspace_Developer.html'
        writeFd = open(filePath, 'w', encoding='utf-8')
        writeFd.write(finalHtml)
        writeFd.close()

        # Now process Chapter 2 table - update with enhanced function
        self.addEnhancedUnreportedColumnToChapter2(filePath, fileStats, totalIssues)

        # Create XML report
        testsuites = ET.Element('testsuites')
        testsuite = ET.SubElement(testsuites, "testsuite")
        testsuite.set("name", "Polyspace验证结果: " + resultText)

        # Total test cases
        if isinstance(redCount, int) and isinstance(orangeCount, int):
            testcaseNum = redCount + orangeCount
        else:
            testcaseNum = "NA"
        if testcaseNum == 0:
            testcaseNum = 1
        testsuite.set("tests", str(testcaseNum))

        # Failed test cases (unreported orange)
        testsuite.set("failures", str(totalIssues))
        # Error test cases (red)
        testsuite.set("errors", str(redCount))
        # Skipped test cases (reported orange)
        testsuite.set("skipped", str(totalReported))

        # Add error testcase elements
        if isinstance(redCount, int):
            for i in range(redCount):
                testcase = ET.SubElement(testsuite, "testcase")
                testcase.set("classname", "红色检查")
                testcase.set("name", f"红色检查{i + 1}")
                testcase.set("time", "0.01")
                skip = ET.SubElement(testcase, "error")
                skip.set("message", "红色检查")

        # Add failure testcase elements
        if isinstance(totalIssues, int):
            for i in range(totalIssues):
                testcase = ET.SubElement(testsuite, "testcase")
                testcase.set("classname", "橙色检查")
                testcase.set("name", f"橙色检查{i + 1}")
                testcase.set("time", "0.01")
                failure = ET.SubElement(testcase, "failure")
                failure.set("message", "橙色检查")

        # Add skipped testcase elements
        if isinstance(totalReported, int):
            for i in range(totalReported):
                testcase = ET.SubElement(testsuite, "testcase")
                testcase.set("classname", "报备检查")
                testcase.set("name", f"报备检查{i + 1}")
                testcase.set("time", "0.01")
                skip = ET.SubElement(testcase, "skipped")
                skip.set("message", "报备检查")

        xmlPath = self.reportPath + '/CATL_Polyspace_Developer.xml'
        # Write XML
        tree = ET.ElementTree(testsuites)
        tree.write(xmlPath, encoding="utf-8", xml_declaration=True, pretty_print=True)

    def addEnhancedUnreportedColumnToChapter2(self, html_file_path, fileStats, totalUnreportedOrange):
        """
        Add enhanced '未报备橙红色项' column to Chapter 2 table with proper formatting

        Changes:
        1. Column name changed from '未报备橙色' to '未报备橙红色项'
        2. Color coding: green for 0, red for >0
        3. Proper column width adjustment
        4. Total calculation at footer
        """
        try:
            # Read HTML file
            with open(html_file_path, 'r', encoding='utf-8') as f:
                html_content = f.read()

            tree = etree.HTML(html_content)

            # Find Chapter 2 table
            target_table = None

            # Strategy 1: Look for specific title
            chapter2_headings = tree.xpath(
                "//h2[.//span[@class='rgSect1TitleText'][contains(., 'Polyspace - CP_Result 的运行时检查摘要')]]"
            )

            for heading in chapter2_headings:
                # Get the first table after this heading
                tables = heading.xpath("./following::table[1]")
                if tables:
                    target_table = tables[0]
                    break

            # Strategy 2: Look for table with specific headers
            if not target_table:
                all_tables = tree.xpath("//table")
                for table in all_tables:
                    headers = table.xpath(".//thead//td//text()")
                    if headers and "文件" in headers and "橙色" in headers:
                        target_table = table
                        break

            if not target_table:
                print("Warning: Could not find Chapter 2 table")
                return False

            # Process the table
            return self._processEnhancedChapter2Table(
                target_table, fileStats, totalUnreportedOrange, tree, html_file_path
            )

        except Exception as e:
            print(f"Error processing Chapter 2 table: {str(e)}")
            import traceback
            traceback.print_exc()
            return False

    def _processEnhancedChapter2Table(self, table, fileStats, totalUnreportedOrange, tree, html_file_path):
        """Process Chapter 2 table with enhanced formatting"""
        try:
            # Get table structure
            header_rows = table.xpath(".//thead/tr")
            if not header_rows:
                print("Warning: No header row found in table")
                return False

            header_row = header_rows[0]
            body_rows = table.xpath(".//tbody/tr")
            footer_rows = table.xpath(".//tfoot/tr")

            # 1. Get current column structure and adjust column widths
            colgroup = table.xpath(".//colgroup")
            if colgroup:
                # Original table has 6 columns, we need to add 1 more
                # Calculate new width: 100% / 7 = 14.2857% each
                colgroup = colgroup[0]

                # Get existing columns
                existing_cols = colgroup.xpath("./col")

                # Adjust all existing columns to new width
                new_width = "14.285714285714285%"  # 100% / 7

                for col in existing_cols:
                    col.set("style", f"width:{new_width};")

                # Add new column for '未报备橙红色项'
                new_col = etree.SubElement(colgroup, "col")
                new_col.set("style", f"width:{new_width};")

            # 2. Add new column header
            # Count current number of header cells
            current_header_cells = header_row.xpath(".//td")

            # Create new header cell
            new_header_cell = etree.SubElement(header_row, "td")
            new_header_p = etree.SubElement(new_header_cell, "p")
            new_header_span = etree.SubElement(new_header_p, "span")
            new_header_span.text = "未报备橙/红色项"
            new_header_span.set('style', 'font-weight:bold;')

            # 3. Process body rows
            for row in body_rows:
                cells = row.xpath(".//td")
                if cells:
                    first_cell_text = ''.join(cells[0].xpath(".//text()")).strip()
                    file_unreported = 0

                    # Match filename to statistics
                    for fileName, stats in fileStats.items():
                        # Try multiple matching strategies
                        if (fileName in first_cell_text or
                                fileName + '.c' == first_cell_text or
                                fileName + '.h' == first_cell_text):
                            file_unreported = stats.get('unreported', 0)
                            break

                    # Add new cell with color coding
                    new_cell = etree.SubElement(row, "td")
                    new_p = etree.SubElement(new_cell, "p")
                    new_span = etree.SubElement(new_p, "span")
                    new_span.text = str(file_unreported)

                    # Color coding: green for 0, red for >0
                    if file_unreported == 0:
                        new_span.set('style', 'color:#008000;')  # Green
                    else:
                        new_span.set('style', 'color:#FF0000; font-weight:bold;')  # Red, bold

            # 4. Process footer row (total row)
            for row in footer_rows:
                # Create new total cell
                new_cell = etree.SubElement(row, "td")
                new_p = etree.SubElement(new_cell, "p")
                new_span = etree.SubElement(new_p, "span")
                new_span.text = str(totalUnreportedOrange)

                # Color coding for total
                if totalUnreportedOrange == 0:
                    new_span.set('style', 'color:#008000; font-weight:bold;')  # Green, bold
                else:
                    new_span.set('style', 'color:#FF0000; font-weight:bold;')  # Red, bold

            # 5. Save changes
            final_html = etree.tostring(tree, method='html', encoding='unicode', with_tail=True)
            with open(html_file_path, 'w', encoding='utf-8') as f:
                f.write(final_html)

            return True

        except Exception as e:
            print(f"Error processing table: {str(e)}")
            import traceback
            traceback.print_exc()
            return False

    def isCheckInRuleList(self, check_type, rawRules):
        """检查检查类型是否在规则列表中"""
        if not rawRules:
            return False

        rawRules_str = str(rawRules).strip()
        if not rawRules_str:
            return False

        # 用分隔符分割规则
        rules = re.split(r'[\\、]', rawRules_str)

        for rule in rules:
            rule = rule.strip()
            if not rule:
                continue

            # 检查是否是复合规则（用&连接）
            if '&' in rule:
                sub_rules = rule.split('&')
                # 检查当前检查类型是否在任何一个子规则中
                for sub_rule in sub_rules:
                    if check_type == sub_rule.strip():
                        return True
            else:
                # 单个规则直接比较
                if check_type == rule:
                    return True

        return False

    def isFileMatchModules(self, file_path, modules):
        """检查文件是否匹配模块"""
        if not modules:
            return False

        modules_str = str(modules).strip()
        if not modules_str or modules_str == 'All':
            return True

        # 用分隔符分割模块
        module_list = re.split(r'[\\、]', modules_str)

        for module_pattern in module_list:
            module_pattern = module_pattern.strip()
            if not module_pattern:
                continue

            # 将通配符模式转换为正则表达式
            regex_pattern = module_pattern.replace('*', '.*').replace('?', '.')

            # 检查文件路径是否匹配
            if re.search(regex_pattern, file_path, re.IGNORECASE):
                return True

        return False

    def getModuleWhiteList(self):
        """从Excel获取模块白名单（第二列）"""
        module_white_list = []

        try:
            wb = load_workbook(self.polyspaceRules)

            # 首先尝试找到合适的sheet
            target_sheet_name = None
            for sheet_name in wb.sheetnames:
                if '报备项' in sheet_name or 'WhiteList' in sheet_name or 'Module' in sheet_name:
                    target_sheet_name = sheet_name
                    break

            if not target_sheet_name:
                # 如果没有找到特定名称的sheet，使用第一个sheet
                target_sheet_name = wb.sheetnames[0]

            ws = wb[target_sheet_name]

            # 读取表头
            firstRow = next(ws.iter_rows(min_row=1, max_row=1))
            headers = [cell.value for cell in firstRow]

            # 查找WhiteList Module列（第二列）
            module_index = -1
            for i, header in enumerate(headers):
                if header and ('WhiteList Module' in str(header) or '模块' in str(header)):
                    module_index = i
                    break

            # 如果没有找到特定列名，假设第二列是模块白名单
            if module_index == -1 and len(headers) >= 2:
                module_index = 1  # 第二列索引

            if module_index >= 0:
                # 读取所有行的模块白名单
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if module_index < len(row) and row[module_index]:
                        modules = str(row[module_index]).strip()
                        if modules:
                            # 按分隔符分割
                            for module in re.split(r'[\\、]', modules):
                                module = module.strip()
                                if module:
                                    module_white_list.append(module)

            wb.close()

        except Exception as e:
            print(f"读取模块白名单Excel失败: {str(e)}")
            import traceback
            traceback.print_exc()

        return module_white_list

    def isFileInModuleWhiteList(self, fileName, moduleWhiteList):
        """检查文件名是否在模块白名单中"""
        if not fileName.lower().endswith('.c'):
            return False

        for pattern in moduleWhiteList:
            try:
                # 将通配符模式转换为正则表达式
                # 注意：Excel中的模式如 "*/FaultM*.c"
                pattern = pattern.strip()
                if not pattern:
                    continue

                # 确保以.c结尾
                if not pattern.endswith('.c'):
                    pattern = pattern + '*.c' if '*' not in pattern else pattern

                # 转换通配符为正则表达式
                regex_pattern = pattern.replace('*', '.*').replace('?', '.')

                # 检查文件名是否匹配
                if re.match(regex_pattern, fileName, re.IGNORECASE):
                    print(f"文件 {fileName} 匹配模块白名单模式: {pattern}")
                    return True
            except Exception as e:
                print(f"检查文件 {fileName} 匹配模式 {pattern} 时出错: {str(e)}")
                continue

        return False

    def addUnreportedOrangeColumnToChapter2(self, html_file_path, fileStats, totalUnreportedOrange):
        """Add unreported orange column to Chapter 2 table - enhanced version"""
        try:
            # Read HTML file
            with open(html_file_path, 'r', encoding='utf-8') as f:
                html_content = f.read()

            tree = etree.HTML(html_content)

            # Find Chapter 2 table
            target_table = None

            # Multiple search strategies
            search_patterns = [
                "//h2[.//span[@class='rgSect1TitleText'][contains(., 'Polyspace - CP_Result 的运行时检查摘要')]]//following::table[1]",
                "//table[.//thead//td[contains(., '文件')] and .//thead//td[contains(., '橙色')]]",
                "//h2[contains(., '第2章')]//following::table[1]"
            ]

            for pattern in search_patterns:
                tables = tree.xpath(pattern)
                if tables:
                    target_table = tables[0]
                    break

            if not target_table:
                print("Warning: Could not find Chapter 2 table")
                # Debug: print all tables for troubleshooting
                all_tables = tree.xpath("//table")
                print(f"Total tables found: {len(all_tables)}")
                for i, table in enumerate(all_tables[:3]):  # First 3 tables
                    headers = table.xpath(".//thead//td//text()")
                    print(f"Table {i + 1} headers: {headers}")
                return False

            # Process the table
            return self._processTableWithEnhancedFormatting(
                target_table, fileStats, totalUnreportedOrange, tree, html_file_path
            )

        except Exception as e:
            print(f"Error in addUnreportedOrangeColumnToChapter2: {str(e)}")
            return False

    def _processTableWithEnhancedFormatting(self, table, fileStats, totalUnreportedOrange, tree, html_file_path):
        """Process table with all enhanced formatting requirements"""
        try:
            # Get table elements
            header_row = table.xpath(".//thead/tr[1]")[0]
            body_rows = table.xpath(".//tbody/tr")
            footer_rows = table.xpath(".//tfoot/tr")

            # 1. Adjust column widths in colgroup
            colgroup = table.xpath(".//colgroup")
            if colgroup:
                colgroup = colgroup[0]
                existing_cols = colgroup.xpath("./col")

                # Calculate new width for 7 columns (100% / 7)
                new_width = "14.285714285714285%"

                # Update existing columns
                for col in existing_cols:
                    col.set("style", f"width:{new_width};")

                # Add new column
                new_col = etree.SubElement(colgroup, "col")
                new_col.set("style", f"width:{new_width};")

            # 2. Add new column header
            new_header_cell = etree.SubElement(header_row, "td")
            new_header_p = etree.SubElement(new_header_cell, "p")
            new_header_span = etree.SubElement(new_header_p, "span")
            new_header_span.text = "未报备橙色项"
            new_header_span.set('style', 'font-weight:bold;')

            # 3. Process body rows
            for row in body_rows:
                cells = row.xpath(".//td")
                if cells:
                    filename_cell = cells[0]
                    filename_text = ''.join(filename_cell.xpath(".//text()")).strip()

                    # Find matching file statistics
                    unreported_count = 0
                    filename_clean = filename_text.replace('.c', '').replace('.h', '')

                    for file_key, stats in fileStats.items():
                        if file_key in filename_clean or filename_clean in file_key:
                            unreported_count = stats.get('unreported', 0)
                            break

                    # Add new cell with conditional formatting
                    new_cell = etree.SubElement(row, "td")
                    new_p = etree.SubElement(new_cell, "p")
                    new_span = etree.SubElement(new_p, "span")
                    new_span.text = str(unreported_count)

                    # Apply color based on value
                    if unreported_count == 0:
                        new_span.set('style', 'color:#008000;')  # Green for zero
                    else:
                        new_span.set('style', 'color:#FF0000; font-weight:bold;')  # Red and bold for non-zero

            # 4. Process footer row (total)
            for row in footer_rows:
                # Add total for new column
                new_cell = etree.SubElement(row, "td")
                new_p = etree.SubElement(new_cell, "p")
                new_span = etree.SubElement(new_p, "span")
                new_span.text = str(totalUnreportedOrange)

                # Apply color to total
                if totalUnreportedOrange == 0:
                    new_span.set('style', 'color:#008000; font-weight:bold;')  # Green, bold
                else:
                    new_span.set('style', 'color:#FF0000; font-weight:bold;')  # Red, bold

            # 5. Save the modified HTML
            final_html = etree.tostring(tree, method='html', encoding='unicode', with_tail=True)
            with open(html_file_path, 'w', encoding='utf-8') as f:
                f.write(final_html)

            print(f"Successfully updated table with '未报备橙红色项' column")
            print(f"- Column name: 未报备橙红色项")
            print(f"- Column width adjusted to 1/7 of table width")
            print(f"- Color coding: Green for 0, Red for >0")
            print(f"- Total value: {totalUnreportedOrange} (color: {'Green' if totalUnreportedOrange == 0 else 'Red'})")

            return True

        except Exception as e:
            print(f"Error in table processing: {str(e)}")
            import traceback
            traceback.print_exc()
            return False

    #Get the Build path
    def getPath(self):
        return Util.getBuildPath()

    # Get path of /Polyspace.psprj
    def getProject(self):
        return os.path.join(self.polyspacePath, 'Polyspace.psprj').replace('\\', '/')

    def setAllConfig(self):
        print("Config all path ")
        # Get project path (already normalized in start())
        self.appPath = getattr(self, 'appPath', Util.normalize_path(Util.getProjectPath()))

        # Get build path (already normalized in start())
        self.buildPath = getattr(self, 'buildPath', Util.normalize_path(Util.getBuildPath()))

        if self.allConfig.get("polyspace") is None:
            Util.reportError("ERROR: Please config file config.ini")

        # Get polyspace output path
        if self.allConfig["polyspace"].get('polyspace_path') is None:
            # Ensure we don't create duplicate Build directory
            base_build_path = self.buildPath
            # If path already ends with Build, don't add it again
            if base_build_path.endswith(os.sep + 'Build') or base_build_path.endswith('/Build'):
                # Remove trailing Build if we're going to add CodeVerify/Polyspace
                polyspace_path = os.path.join(base_build_path, 'CodeVerify', 'Polyspace')
            else:
                polyspace_path = os.path.join(base_build_path, 'CodeVerify', 'Polyspace')

            self.allConfig["polyspace"]['polyspace_path'] = polyspace_path

        self.polyspacePath = Util.normalize_path(self.allConfig['polyspace']['polyspace_path'])

        # Handle relative paths
        if '..' in self.polyspacePath:
            config_dir = os.path.dirname(Util.getConfigPath())
            self.polyspacePath = Util.normalize_path(os.path.join(config_dir, self.polyspacePath))

        # Ensure the path exists
        if not os.path.exists(self.polyspacePath):
            os.makedirs(self.polyspacePath, exist_ok=True)

        # Get template file path
        template_path = os.path.join(self.buildPath, 'Tools', 'Components', 'Polyspace', 'templates',
                                     'polyspace_template.psprj')
        if not os.path.exists(template_path):
            template_path = os.path.join(self.buildPath, 'tools', 'Components', 'Polyspace', 'templates',
                                         'polyspace_template.psprj')
            if not os.path.exists(template_path):
                Util.reportError(f"Template file not found: {template_path}")

        self.templatePath = template_path

        # Get polyspace config file
        self.polyspaceCheck = os.path.join(self.buildPath, "VerifyCfg", "polyspace_check.ini")
        if not os.path.exists(self.polyspaceCheck):
            Util.reportError(f"polyspace_check.ini not found at: {self.polyspaceCheck}")

        # Result path after script execution
        self.launchResult = os.path.join(self.polyspacePath, 'Component', 'CP_Result')
        if (not os.path.exists(self.launchResult)):
            os.makedirs(self.launchResult, exist_ok=True)

        # Report storage path
        self.reportPath = os.path.join(self.polyspacePath, 'Component', 'CP_Result', 'Polyspace-Doc')
        if (not os.path.exists(self.reportPath)):
            os.makedirs(self.reportPath, exist_ok=True)

        self.generateLaunch = os.path.join(self.polyspacePath, 'generateLaunch.bat')

        # Get matlab path
        if (os.environ.get("CI_POLYSPACE_BASE", default=None) is not None):
            self.matlabPath = os.environ.get("CI_POLYSPACE_BASE", default=None)
        elif (self.allConfig['polyspace'].get('matlab_path') is not None):
            self.matlabPath = self.allConfig['polyspace']['matlab_path']
            print("Find polyspace path from config.ini: " + self.matlabPath)
        else:
            print(Fore.RED + "ERROR: Please config polyspace environment in system environment or config.ini")
            exit(-1)

        # Verify polyspace executable
        proverPath = self.matlabPath + '/polyspace/bin/polyspace-code-prover-server.exe'
        if (not os.path.exists(proverPath)):
            proverPath = self.matlabPath + '/polyspace/bin/polyspace-code-prover.exe'
            if (not os.path.exists(proverPath)):
                Util.reportError(proverPath + ' don\'t exist! Please check polyspace config in config.ini')
        # Check version
        polyspaceVersion = (subprocess.run(proverPath + ' -v', capture_output=True))
        if (polyspaceVersion.stdout.decode(encoding='utf-8').find('2023b') == -1):
            Util.reportError('Your matlab is not the version of 2023b.')

        self.compiler = self.allConfig['polyspace']['compiler']
        self.target = self.allConfig['polyspace']['target']
        self.psprjPath = os.path.join(self.polyspacePath, 'polyspace.psprj')
        self.componentPath = self.polyspacePath + '\\Component'
        self.problemSet = set()
        self.allFile = set()
        self.noSuchFile = set()
        self.allNumber = 0
        self.polyspaceRules = os.path.join(self.buildPath, "Tools", "Components", "Polyspace", "doc",
                                           "Polyspace_Rule.xlsx")

    def getTemplate(self):
        with open(self.templatePath, 'r', encoding='utf-8') as f:
            data = f.read()
        return data

    # Get all the configuration
    def initConfig(self, path):
        print("Init all the Configration ")
        self.allNumber = 0
        try:
            config = configparser.ConfigParser()
            self.allConfig = {}
            print("Initialize configuration and read .ini configuration file")
            config.read(path, encoding='utf-8')
            # Partition according to []
            seclist = config.sections()
            for i in seclist:
                self.allConfig[i] = {}
                data = config.items(i)
                for j in data:
                    self.allConfig[i][j[0]] = j[1]
        except Exception as e:
            print("ERROR:  Failed to read configuration file")
            raise e

    def getCpu(self):
        maxCpuNmuber = os.getenv('CI_POLYSPACE_MAX_CPU_NUMBER')
        if maxCpuNmuber is not None:
            try:
                if int(maxCpuNmuber) >= 128:
                    return 128
                return int(maxCpuNmuber)
            except ValueError:
                print("Warning: maxCpuNmuber must be an integer, using CPU count instead")

        if multiprocessing.cpu_count() >= 128:
            return 128
        return multiprocessing.cpu_count()

    def runLaunchScript(self):
        if(not os.path.exists(self.polyspacePath+'\\Component\\CP_Result\\.settings')):
            os.makedirs(self.polyspacePath+'\\Component\\CP_Result\\.settings')

        self.fileMove(self.polyspacePath+'\\Component\\CP_Result\\.settings\\.polyspace_conf.psprj', self.polyspacePath+\
            '\\Polyspace\\.polyspace_conf.psprj',self.polyspacePath+'\\Component\\CP_Result\\.settings')

        self.fileMove(self.polyspacePath+'\\Component\\CP_Result\\.settings\\launchingCommand.bat', self.polyspacePath+\
            '\\Polyspace\\launchingCommand.bat',self.polyspacePath+'\\Component\\CP_Result\\.settings')

        self.fileMove(self.polyspacePath+'\\Component\\CP_Result\\.settings\\options_command.txt', self.polyspacePath+\
            '\\Polyspace\\options_command.txt',self.polyspacePath+'\\Component\\CP_Result\\.settings')

        self.fileMove(self.polyspacePath+'\\Component\\CP_Result\\.settings\\source_command.txt', self.polyspacePath+\
            '\\Polyspace\\source_command.txt',self.polyspacePath+'\\Component\\CP_Result\\.settings')

        launchPath = self.polyspacePath+'\\Component\\CP_Result\\.settings\\launchingCommand.bat'
        fd = open(launchPath, 'r', encoding='utf-8')
        text = fd.read()
        fd.close()
        print(Fore.GREEN + text)

        fd = open(launchPath, 'w+', encoding='utf-8')
        rep = re.compile('-options-file( ".*?")').findall(text)[0]
        # update launchingCommand.bat
        text = text.replace(rep, ' "'+self.polyspacePath+'\\Component\\CP_Result\\.settings\\options_command.txt"')

        rep = re.compile('-results-dir(.*?)%\*').findall(text)[0]
        text = text.replace(rep, ' "'+self.polyspacePath+'\\Component\\CP_Result" ')
        fd.write(text)
        fd.close()

        #update option-command.txt
        optionCommandPath = self.polyspacePath+'\\Component\\CP_Result\\.settings\\options_command.txt'
        ocfd = open(optionCommandPath, 'r', encoding='utf-8')
        lines = ocfd.readlines()
        ocfd.close()

        ocfd = open(optionCommandPath, 'w+', encoding='utf-8')
        for line in lines:
            if(re.compile('-sources-list-file(.*?)').findall(line) != []):
                line = '-sources-list-file '+self.polyspacePath+'\\Component\\CP_Result\\.settings\\source_command.txt\n'
                ocfd.write(line)
            else:
                ocfd.write(line)
        ocfd.close()
        if(subprocess.run(launchPath,shell=True).returncode != 0):
            Util.reportError(f'Run launch script Failed')
            raise ValueError
        self.clean(1)

    def runReportScript(self):
        reportScript = '''
"{reportGene}" -template {matlabPath}\\toolbox\\polyspace\\psrptgen\\templates\\Developer.rpt -results-dir {launchResult} -output-name {reportPath}\\Developer -format doc
echo ..............................................................................
 '''

        reportScript = reportScript.format(reportGene=self.matlabPath+\
            '\\polyspace\\bin\\polyspace-report-generator.exe', \
            matlabPath = self.matlabPath, reportPath = self.reportPath, \
            launchResult=self.launchResult)
        reportFd = open(self.polyspacePath+'/report.bat', 'w+', encoding='utf-8')
        print(Fore.GREEN + reportScript)
        reportFd.write(reportScript)
        reportFd.close()
        subprocess.call(self.polyspacePath+'/report.bat')

        if(os.path.exists(self.polyspacePath+'/Report/report.bat')):
            os.remove(self.polyspacePath+'/Report/report.bat')
        shutil.move(self.polyspacePath+'/report.bat', self.polyspacePath+'\\Report')

    # Generate PSPRJ file in psprjPath
    def generatePSPRJFile(self):
        print(Fore.GREEN + "Generate PSPRJ file in "+self.polyspacePath)

        # Ensure directory exists
        os.makedirs(self.polyspacePath, exist_ok=True)

        # Get the template file
        template_path = self.templatePath

        if not os.path.exists(template_path):
            Util.reportError(f"Template file not found: {template_path}")

        with open(template_path, 'r', encoding='utf-8') as f:
            template = f.read()

        (com, tar) = self.getComTag()

        # Read environment variables to obtain chip template information
        self.getCfgEnv()

        # Get each section content
        cfile_content = self.getCFile()
        include_content = self.readInclude()

        # Remove any test file logic - just use the actual source files
        if not cfile_content or cfile_content.strip() == "":
            print(Fore.RED + "ERROR: No source file content generated!")
            print("This will cause 'No available source files' error")
            exit(-1)

        # Generate complete PSPRJ content
        psprj_content = template.format(
            author=getpass.getuser(),
            date=datetime.datetime.now().strftime("%d/%m/%Y"),
            project=self.getProject(),
            cfile=cfile_content,
            include=include_content,
            component=self.getResultPath(),
            ansfile=self.getCfileAnya(),
            cpu=self.getCpu(),
            eleD=self.getConfig(),
            eleI=self.readIncludeNoNum(),
            compiler=com,
            target=tar
        ).replace("\\", '/')

        psprj_file = os.path.join(self.polyspacePath, 'Polyspace.psprj')

        with open(psprj_file, 'w', encoding='utf-8') as f:
            f.write(psprj_content)

        # Verify file
        file_size = os.path.getsize(psprj_file)

        # Count source files
        source_count = psprj_content.count('<file path=')

        if source_count == 0:
            print(Fore.RED + "ERROR: No source files found in PSPRJ!")
            print("This will cause 'No available source files' error")
            exit(-1)


    def getCfgEnv(self):
        data = []
        try:
            if (compilerEnv == 'cmake'):
                env_cmake_path = os.path.join(self.buildPath, 'Cmake', 'env.cmake')

                if not os.path.exists(env_cmake_path):
                    # Try other possible locations
                    possible_paths = [
                        env_cmake_path,
                        os.path.join(self.buildPath, 'env.cmake'),
                        os.path.join(Util.getProjectPath(), 'Build', 'Cmake', 'env.cmake'),
                    ]

                    for path in possible_paths:
                        if os.path.exists(path):
                            env_cmake_path = path
                            break

                if not os.path.exists(env_cmake_path):
                    Util.reportError(f"env.cmake not found at: {env_cmake_path}")

                with open(env_cmake_path, 'r', encoding='utf-8') as f:
                    data = f.read()

                # Get Compiler     greenhills_arm
                self.buildTool = re.compile('set\(COMPILER (.*?)\)').findall(data)[0]
                # Get CHIP
                self.chipTool = re.compile('set\(CHIP_NAME (.*?)\)').findall(data)[0]
                # Get projectName
                self.projectName = re.compile('set\(PROJECT_NAME (.*?)\)').findall(data)[0]
                # Get compiler Path
                if (os.environ.get("CI_COMPILER_BASE", default=None) is not None):
                    self.compilerBase = os.environ.get("CI_COMPILER_BASE", default=None)
                    print(Fore.GREEN + "Compiler: " + self.buildTool + '\nCHIP: ' + \
                          self.chipTool + '\nCI_COMPILER_BASE: ' + self.compilerBase + '\nPROJECT_NAME: ' + self.projectName)
                else:
                    self.compilerBase = re.compile('set\(COMPILER_BASE "(.*?)"\)').findall(data)[0]
                    print(Fore.GREEN + "Compiler: " + self.buildTool + '\nCHIP: ' + \
                          self.chipTool + '\nCOMPILER_BASE: ' + self.compilerBase + '\nPROJECT_NAME: ' + self.projectName)

            elif (compilerEnv == 'emake'):
                project_ini_path = os.path.join(self.buildPath, 'cfg', 'project.ini')

                if not os.path.exists(project_ini_path):
                    possible_paths = [
                        project_ini_path,
                        os.path.join(self.buildPath, 'project.ini'),
                        os.path.join(Util.getProjectPath(), 'Build', 'cfg', 'project.ini'),
                    ]

                    for path in possible_paths:
                        if os.path.exists(path):
                            project_ini_path = path
                            break

                if not os.path.exists(project_ini_path):
                    Util.reportError(f"project.ini not found at: {project_ini_path}")

                with open(project_ini_path, 'r', encoding='utf-8') as f:
                    data = f.read()

                # Get Compiler
                build_toolchain_match = re.compile('build_toolchain\s*=\s*(.*)').findall(data)
                if build_toolchain_match:
                    self.buildTool = build_toolchain_match[0]
                else:
                    self.buildTool = 'unknown'

                # Get CHIP
                chip_match = re.compile('chip\s*=\s*(.*)').findall(data)
                if chip_match:
                    self.chipTool = chip_match[0]
                else:
                    self.chipTool = 'unknown'

                # Get projectName
                project_name_match = re.compile('project_name\s*=\s*(.*)').findall(data)
                if project_name_match:
                    self.projectName = project_name_match[0]
                else:
                    self.projectName = 'unknown'

                # Get compiler Path
                if (os.environ.get("CI_COMPILER_BASE", default=None) is not None):
                    self.compilerBase = os.environ.get("CI_COMPILER_BASE", default=None)
                    print(Fore.GREEN + f"build_toolchain: {self.buildTool}\nchip: {self.chipTool}" + \
                          f"\nCI_COMPILER_BASE: {self.compilerBase}\nproject_name: {self.projectName}")
                else:
                    compiler_path_match = re.compile('compiler_path\s*=\s*(.*)').findall(data)
                    if compiler_path_match:
                        self.compilerBase = compiler_path_match[0]
                    else:
                        self.compilerBase = 'unknown'
                    print(Fore.GREEN + f"build_toolchain: {self.buildTool}\nchip: {self.chipTool}" + \
                          f"\nCOMPILER_BASE: {self.compilerBase}\nproject_name: {self.projectName}")
        except FileNotFoundError:
            Util.reportError(f"compilerEnv: {compilerEnv}, env.cmake or cfg/project.ini not found")
        except Exception as e:
            Util.reportError(
                f"Error reading cfg env: {str(e)}, {type(e).__name__} at line {sys.exc_info()[-1].tb_lineno}")

    def getsourcefile(self):
        lineKey = False
        data = []
        suppressPath = []
        defaultKey = []

        # Read all Rte*.c file paths (for full mode)
        rte_files = []

        # In full mode, read all Rte*.c files from source.cmake
        if polyspaceAnalysisMode == 'full':
            source_cmake_path = os.path.join(self.buildPath, 'Cmake', 'source.cmake')
            if os.path.exists(source_cmake_path):
                try:
                    with open(source_cmake_path, 'r', encoding='utf-8') as f:
                        content = f.read()

                    import re
                    # Find source settings
                    pattern = r'set\(source(.*?)\)'
                    match = re.search(pattern, content, re.S | re.M)

                    if not match:
                        patterns = [
                            r'set\(\s*source(.*?)\)',
                            r'set\(SOURCES?(.*?)\)',
                            r'set\(\s*SOURCES?(.*?)\)',
                        ]

                        for pattern in patterns:
                            match = re.search(pattern, content, re.S | re.M)
                            if match:
                                break

                    if match:
                        source_content = match.group(1)
                        raw_lines = source_content.split('\n')
                        source_files = [i.replace('\\', '/').strip() for i in raw_lines if i.strip()]

                        # Filter Rte*.c files
                        for source_line in source_files:
                            if source_line and not source_line.startswith('#'):
                                # Resolve relative path
                                if source_line.startswith('./'):
                                    resolved_path = Util.resolveRelativePath(source_line[2:])
                                    filename = os.path.basename(resolved_path)

                                    # Check if it's Rte*.c file (case-insensitive)
                                    if re.match(r'^Rte.*\.c$', filename, re.IGNORECASE):
                                        rte_files.append(resolved_path)
                except Exception:
                    # Silent error handling
                    pass

        source = self.readSource().split('\n\t')
        include = self.readInclude().split('\n\t')
        sourceKey = []
        suffixesSource = ['_Cfg.c', '_Lcfg.c', '_Cbk.c', '_PBcfg.c', '_define.c', '_Callout.c', '_Callout_Stubs.c',
                          '_Hal_Core.c', '_Irq.c', '.c']
        suffixesInclude = ['_Cfg.h', '_Lcfg.h', '_Cbk.h', '_PBcfg.h', '_define.h', '_Callout.h', '_Callout_Stubs.h',
                           '_Hal_Core.h', '_GeneralTypes.h', '_Type.h', '_Types.h', '_Memmap.h', '_Int.h', '_Irq.h',
                           '.h']
        prefixesKey = ['SchM_', 'Rte_']
        suffixesSourceLower = [s.lower() for s in suffixesSource]
        suffixesIncludeLower = [s.lower() for s in suffixesInclude]
        prefixesKeyLower = [s.lower() for s in prefixesKey]

        # Read polyspace_check.ini configuration
        with open(self.polyspaceCheck, 'r', encoding='utf-8') as f:
            for line in f.readlines():
                line = line.replace("\\", "/")
                if not lineKey:
                    # Read SpecifyAnalysis section for specify mode
                    if line.startswith('[SpecifyAnalysis]') and polyspaceAnalysisMode == 'specify':
                        lineKey = True
                        continue
                    # Read SuppressAnalysis section for full and guard modes
                    if line.startswith('[SuppressAnalysis]') and (
                            polyspaceAnalysisMode == 'full' or polyspaceAnalysisMode == 'guard'):
                        lineKey = True
                        continue
                elif lineKey == True and line.startswith(r'['):
                    break
                else:
                    if not line.startswith('./'):
                        continue
                    if (polyspaceAnalysisMode == "guard") or (polyspaceAnalysisMode == 'full'):
                        suppressPath.append(line[1:])  # Remove leading '.' for path matching
                    else:
                        # For specify mode, add file directly
                        resolved_path = Util.resolveRelativePath(line.replace("\n", '').replace('./', ''))
                        data.append([os.path.dirname(resolved_path), os.path.basename(resolved_path)])

        # Add default analysis files (only for guard mode)
        lineKey = False
        if polyspaceAnalysisMode == "guard":
            with open(self.polyspaceCheck, 'r', encoding='utf-8') as f:
                for line in f.readlines():
                    line = line.replace("\\", "/")
                    if not lineKey:
                        if line.startswith('[DefaultAnalysis]'):
                            lineKey = True
                            continue
                    elif lineKey == True and line.startswith(r'['):
                        break
                    else:
                        if not line.startswith('./'):
                            continue
                        else:
                            resolved_path = Util.resolveRelativePath(line.replace("\n", '').replace('./', ''))
                            data.append([os.path.dirname(resolved_path), os.path.basename(resolved_path)])

            # Guard mode needs to process files in changedFiles.txt
            if polyspacechangeFilePath and os.path.exists(polyspacechangeFilePath):
                with open(polyspacechangeFilePath, 'r', encoding='utf-8') as f:
                    lines = f.readlines()
                    for line in lines:
                        line = line.strip()
                        if not line:
                            continue  # Skip empty lines

                        # Handle multiple formats
                        item = None

                        if line.startswith('./'):
                            item = line[2:]  # Remove "./" prefix

                        elif 'Path:' in line:
                            parts = line.split(',')
                            for part in parts:
                                if 'Path:' in part:
                                    item = part.split('Path:')[1].strip()
                                    if item.startswith('./'):
                                        item = item[2:]
                                    break

                        elif ',' in line:
                            # May be CSV but without explicit Path: label
                            first_part = line.split(',')[0].strip()
                            if first_part.startswith('./'):
                                item = first_part[2:]

                        if not item:
                            # If not in the above formats, try to process as a path
                            item = line.strip()
                            if item.startswith('./'):
                                item = item[2:]

                        if item:
                            # Check if file is in suppress list
                            suppressEnable = False
                            for suppress in suppressPath:
                                suppress = suppress.rstrip('\n')
                                if suppress.lower() in item.lower():
                                    suppressEnable = True
                                    break

                            if not suppressEnable:
                                # Resolve file path
                                filePath = Util.resolveRelativePath(item)
                                if not os.path.exists(filePath):
                                    # Try other possible paths
                                    possible_paths = [
                                        filePath,
                                        os.path.join(Util.getProjectPath(), item),
                                        os.path.join(Util.getSourcePath(),
                                                     item[10:] if item.startswith('SourceCode/') else item),
                                        os.path.join(Util.getProjectPath(), 'Customer', item),
                                    ]

                                    found = False
                                    for alt_path in possible_paths:
                                        if os.path.exists(alt_path):
                                            filePath = alt_path
                                            found = True
                                            break

                                    if not found:
                                        continue  # Skip if file doesn't exist

                                # Check if file has already been added
                                already_added = False
                                for existing_file in data:
                                    if os.path.basename(filePath) == existing_file[1] and os.path.dirname(filePath) == \
                                            existing_file[0]:
                                        already_added = True
                                        break

                                if not already_added:
                                    data.append([os.path.dirname(filePath), os.path.basename(filePath)])

                                    # If it's a .c file, add to sourceKey for pattern matching
                                    if item.lower().endswith('.c'):
                                        filename_base = os.path.basename(item).lower()
                                        for suffix in suffixesSourceLower:
                                            if filename_base.endswith(suffix):
                                                source_item = filename_base[:-len(suffix)]

                                                # Filter prefix
                                                for prefix in prefixesKeyLower:
                                                    if source_item.startswith(prefix):
                                                        source_item = source_item[len(prefix):]
                                                        break

                                                # Add source key
                                                if source_item and source_item not in sourceKey:
                                                    sourceKey.append(source_item)
                                                break

        # Add matching files for guard or full mode
        if (polyspaceAnalysisMode == "guard") or (polyspaceAnalysisMode == 'full'):
            for sourceLine in source:
                # Filter invalid rows
                if not sourceLine or not sourceLine.startswith('./'):
                    continue

                suppressEnable = False
                for suppress in suppressPath:
                    suppress = suppress.rstrip('\n')

                    # Critical check: skip if source file path contains suppress path
                    if suppress.lower() in sourceLine.rstrip('\n').lower():
                        suppressEnable = True
                        break

                # SPECIAL HANDLING: In full mode, force include Rte*.c files even if suppressed
                resolved_path = Util.resolveRelativePath(sourceLine.replace("\n", '').replace('./', ''))
                filename = os.path.basename(resolved_path)
                is_rte_file = False
                if polyspaceAnalysisMode == 'full':
                    is_rte_file = re.match(r'^Rte.*\.c$', filename, re.IGNORECASE)

                # Only add file if not suppressed (or is Rte file in full mode)
                if not suppressEnable or (polyspaceAnalysisMode == 'full' and is_rte_file):
                    if (polyspaceAnalysisMode == "guard"):
                        sourceItem = sourceLine.rstrip('\n').rsplit("/", 1)[1].lower()
                        for key in sourceKey:
                            findKey = False
                            for suffix in suffixesSourceLower:
                                if (key + suffix) == sourceItem:
                                    findKey = True
                                else:
                                    # Filter prefix
                                    for prefix in prefixesKeyLower:
                                        if (prefix + key + suffix) == sourceItem:
                                            findKey = True
                                            break

                                if findKey:
                                    # Check if already added
                                    already_added = False
                                    for existing_file in data:
                                        if os.path.basename(resolved_path) == existing_file[1] and os.path.dirname(
                                                resolved_path) == existing_file[0]:
                                            already_added = True
                                            break

                                    if not already_added:
                                        data.append([os.path.dirname(resolved_path), os.path.basename(resolved_path)])
                                    break

                    elif (polyspaceAnalysisMode == 'full'):
                        # Check if already added
                        already_added = False
                        for existing_file in data:
                            if os.path.basename(resolved_path) == existing_file[1] and os.path.dirname(resolved_path) == \
                                    existing_file[0]:
                                already_added = True
                                break

                        if not already_added:
                            data.append([os.path.dirname(resolved_path), os.path.basename(resolved_path)])

        # SPECIAL HANDLING: In full mode, add all Rte*.c files from source.cmake
        # This ensures all Rte files are included even if not in source list
        if polyspaceAnalysisMode == 'full' and rte_files:
            for rte_file in rte_files:
                if os.path.exists(rte_file):
                    # Check if already in data list
                    already_included = False
                    for item in data:
                        if os.path.basename(item[1]).lower() == os.path.basename(rte_file).lower():
                            already_included = True
                            break

                    if not already_included:
                        data.append([os.path.dirname(rte_file), os.path.basename(rte_file)])

        return data

    def getCFile(self):
        tem = """<root_folder path="file:/{dirname}" isnested="true" filter=".cpp,.cxx,.cc,.c">
        {cfile}
    </root_folder>"""
        data = self.getsourcefile()

        fileList = {}
        for i in data:
            dir_path = i[0].replace('\\', '/')
            file_name = i[1]

            if fileList.get(dir_path) == None:
                fileList[dir_path] = [file_name]
            else:
                fileList[dir_path].append(file_name)

        all = []

        for dir_path, files in fileList.items():
            cfile = '\n'.join(['<file path="{}"/>'.format(j) for j in files])
            all.append(tem.format(dirname=dir_path + '/', cfile=cfile))

        result = '\n'.join(all)
        return result

    # polyspace -generate-launching-script-for Polyspace.psprj
    def getOptionsCommand(self):
        print(Fore.GREEN + "generate -generate-launching-script ")

        polyspace_exe = f'"{self.matlabPath}\\polyspace\\bin\\polyspace.exe"'

        # Determine PSPRJ file path
        psprj_file = os.path.join(self.polyspacePath, 'Polyspace.psprj')

        # Check if PSPRJ file exists
        if not os.path.exists(psprj_file):
            Util.reportError(f"PSPRJ file not found at: {psprj_file}")

        # Switch to correct directory to execute command
        polyspace_dir = os.path.dirname(psprj_file)
        psprj_filename = os.path.basename(psprj_file)

        current_dir = os.getcwd()

        # Build command
        generateLaunchCmd = f'{polyspace_exe} -generate-launching-script-for "{psprj_filename}"'

        # Create generation script
        genLaunchScr = os.path.join(polyspace_dir, 'generate_launching_script.bat')

        # Write script file
        with open(genLaunchScr, 'w+', encoding='utf-8') as GLFd:
            GLFd.write(f'''@echo off
    chcp 65001 >nul
    cd /d "{polyspace_dir}"
    {generateLaunchCmd}
    if %errorlevel% neq 0 (
        echo ERROR: Failed to generate launching script
        exit /b %errorlevel%
    )
    echo SUCCESS: Command completed successfully
    ''')

        print(f"Created script at: {genLaunchScr}")

        try:
            # Switch to polyspace directory to execute command
            os.chdir(polyspace_dir)

            # Execute command
            result = subprocess.run(
                generateLaunchCmd,
                shell=True,
                capture_output=True,
                text=True,
                encoding='utf-8',
                errors='ignore'
            )

            if result.returncode != 0:
                if result.stderr:
                    print(Fore.RED + f"Error: {result.stderr}")
                Util.reportError(f'run -generate-launching-script-for command failed')

        except Exception as e:
            print(Fore.RED + f"Exception during command execution: {e}")
            Util.reportError(f'run -generate-launching-script-for command failed')
        finally:
            # Switch back to original directory
            os.chdir(current_dir)

    def getResultPath(self):
        print(Fore.GREEN + 'Get the result storage path')
        return self.polyspacePath + '/Component/'

    def getCfileAnya(self):
        print(Fore.GREEN + 'Get a list of C files')
        tem = """<root_folder path="file:/{dirname}" isnested="true">
      {cfile}
    </root_folder>"""
        data = self.getsourcefile()

        fileList = {}
        for i in data:
            if fileList.get(i[0]) == None:
                fileList[i[0]] = [i[1]]
            else:
                fileList[i[0]].append(i[1])
        all = []

        for i in fileList.items():
            cfile = '\n'.join(['<file path="{}"/>'.format(j) for j in i[1]])
            all.append(tem.format(dirname=i[0].replace('\\', '/') + '/', cfile=cfile))
        return '\n'.join(all)

    def readIncludeNoNum(self):
        print(Fore.GREEN + 'Read header file')
        tem = '<element>file:/{path}</element>'

        data = []
        if (compilerEnv == 'cmake'):
            include_cmake_path = os.path.join(self.buildPath, 'Cmake', 'include.cmake')

            if not os.path.exists(include_cmake_path):
                Util.reportError(f"include.cmake not found at: {include_cmake_path}")

            with open(include_cmake_path, 'r', encoding='utf-8') as f:
                content = f.read()

            import re
            pattern = r'set\(include_dir(.*?)\)'
            match = re.search(pattern, content, re.S | re.M)

            if not match:
                patterns = [
                    r'set\(\s*include_dir(.*?)\)',
                    r'set\(INCLUDE_DIRS?\s*\((.*?)\)',
                    r'include_directories\s*\((.*?)\)',
                ]

                for pattern in patterns:
                    match = re.search(pattern, content, re.S | re.M)
                    if match:
                        break

            if not match:
                Util.reportError("Could not find set(include_dir ... ) in include.cmake")

            include_content = match.group(1)
            raw_lines = include_content.split('\n')
            data = [i.replace('\\', '/') for i in raw_lines if i.replace(' ', '') != '']

        elif (compilerEnv == 'emake'):
            include_dir_path = os.path.join(self.buildPath, 'cfg', 'file', 'include_dir.txt')

            if not os.path.exists(include_dir_path):
                print(Fore.YELLOW + f"Warning: include_dir.txt not found at: {include_dir_path}")
                # Add default include directories for emake
                data = [
                    './',
                    './inc',
                    './include',
                    './BSW',
                    './APP',
                    './ASW',
                ]
            else:
                with open(include_dir_path, 'r', encoding='utf-8') as f:
                    raw_lines = f.read().split('\n')
                    data = [i.replace('\\', '/') for i in raw_lines if i.replace(' ', '') != '']

        all = []
        self.includeFlag = True
        self.noSuchInclude = set()

        for i in enumerate(data):
            line = i[1]
            if line.startswith('#'):
                continue
            elif str(line).startswith("./"):
                resolved_path = Util.resolveRelativePath(line.strip())
                all.append(tem.format(path=resolved_path.replace('\\', '/') + '/'))
                if not os.path.exists(resolved_path):
                    self.includeFlag = False
                    self.noSuchInclude.add(resolved_path)
            else:
                all.append(tem.format(path=line.strip().replace('\\', '/') + "/"))

        # Add compiler include directories
        if (compilerEnv == 'cmake'):
            toolchain_path = os.path.join(self.buildPath, "Tools", "Components", "Toolchains",
                                          f"{self.buildTool}.cmake")

            if os.path.exists(toolchain_path):
                with open(toolchain_path, 'r', encoding='utf-8') as f:
                    toolchain_content = f.read()

                raw = re.compile('set\(COMPILER_INC  "(.*?)"\)', re.DOTALL).findall(toolchain_content)[0].replace(' ',
                                                                                                                  '')
                allTem = raw.split('\n')
                for index in allTem:
                    if index.strip():
                        resolved_inc = index.replace('${COMPILER_BASE}', self.compilerBase).strip()
                        all.append(tem.format(path=resolved_inc.replace('\\', '/') + "/"))
        elif (compilerEnv == 'emake'):
            if hasattr(self, 'compilerBase') and self.compilerBase:
                compilerPath = self.compilerBase.strip() + "/ctc/include" + "/"
                compilerPath = compilerPath.replace("/ctc/ctc", "/ctc")
                all.append(tem.format(path=compilerPath))

        return "\n\t\t".join(all)

    def getComTag(self):
        self.compiler = self.allConfig['polyspace']['compiler']
        self.target = self.allConfig['polyspace']['target']
        return (self.compiler, self.target)

    # read all include to the polyspace.psprj file
    def readInclude(self):
        print(Fore.GREEN + "Read Header File")
        tem = '<file path="file:/{path}" order="{num}"/>'

        data = []
        if (compilerEnv == 'cmake'):
            include_cmake_path = os.path.join(self.buildPath, 'Cmake', 'include.cmake')

            if not os.path.exists(include_cmake_path):
                print(Fore.RED + f"include.cmake not found at: {include_cmake_path}")
                return "\n\t"

            try:
                with open(include_cmake_path, 'r', encoding='utf-8') as f:
                    content = f.read()

                import re

                pattern = r'set\(include_dir(.*?)\)'
                match = re.search(pattern, content, re.S | re.M)

                if not match:
                    patterns = [
                        r'set\(\s*include_dir(.*?)\)',
                        r'set\(INCLUDE_DIRS?\s*\((.*?)\)',
                        r'set\(\s*INCLUDE_DIRS?\s*\((.*?)\)',
                        r'include_directories\s*\((.*?)\)',
                    ]

                    for pattern in patterns:
                        match = re.search(pattern, content, re.S | re.M)
                        if match:
                            break

                if not match:
                    Util.reportError("Could not find set(include_dir ... ) in include.cmake")

                include_content = match.group(1)
                raw_lines = include_content.split('\n')
                data = [i.replace('\\', '/') for i in raw_lines if i.replace(' ', '') != '']

            except Exception as e:
                return "\n\t"

        elif (compilerEnv == 'emake'):
            # For emake, we need to read from include_dir.txt
            include_dir_path = os.path.join(self.buildPath, 'cfg', 'file', 'include_dir.txt')

            if not os.path.exists(include_dir_path):
                # Try other possible locations
                possible_paths = [
                    include_dir_path,
                    os.path.join(self.buildPath, 'include_dir.txt'),
                    os.path.join(Util.getProjectPath(), 'Build', 'cfg', 'file', 'include_dir.txt'),
                    os.path.join(Util.getProjectPath(), 'cfg', 'file', 'include_dir.txt'),
                ]

                for path in possible_paths:
                    if os.path.exists(path):
                        include_dir_path = path
                        print(f"Found include_dir.txt at: {path}")
                        break

            if not os.path.exists(include_dir_path):
                # Add some default include directories for emake
                data = [
                    './',
                    './inc',
                    './include',
                    './BSW',
                    './APP',
                    './ASW',
                ]
            else:
                with open(include_dir_path, 'r', encoding='utf-8') as f:
                    raw_lines = f.read().split('\n')
                    for line in raw_lines:
                        line = line.strip()
                        if line and not line.startswith('#'):
                            data.append(line.replace('\\', '/'))

        all = []
        for i in enumerate(data):
            line = i[1]
            if line.startswith('#'):
                continue
            elif str(line).startswith("./"):
                resolved_path = Util.resolveRelativePath(line.strip())
                all.append(tem.format(path=resolved_path.replace('\\', '/') + '/', num=i[0]))
            else:
                all.append(tem.format(path=line.strip().replace('\\', '/') + "/", num=i[0]))

        if (compilerEnv == 'cmake'):
            toolchain_path = os.path.join(self.buildPath, "Tools", "Components", "Toolchains",
                                          f"{self.buildTool}.cmake")

            if os.path.exists(toolchain_path):
                with open(toolchain_path, 'r', encoding='utf-8') as f:
                    toolchain_content = f.read()

                raw = re.compile('set\(COMPILER_INC  "(.*?)"\)', re.S).findall(toolchain_content)[0].replace(' ', '')
                allTem = raw.split('\n')
                number = len(data)
                for index in allTem:
                    if index.strip():
                        resolved_inc = index.replace('${COMPILER_BASE}', self.compilerBase)
                        all.append(tem.format(path=resolved_inc.replace('\\', '/') + "/", num=number))
                        number += 1
        elif (compilerEnv == 'emake'):
            number = len(data)
            # Add compiler include directories
            if hasattr(self, 'compilerBase') and self.compilerBase:
                compiler_include = self.compilerBase.strip() + "/ctc/include/"
                compiler_include = compiler_include.replace("/ctc/ctc", "/ctc")
                all.append(tem.format(path=compiler_include, num=number))
                number += 1

        return "\n\t".join(all)

    # read all source to the polyspace.psprj file
    def readSource(self):
        data = []
        print(Fore.GREEN + "Read Source File")

        if (compilerEnv == 'cmake'):
            source_cmake_path = os.path.join(self.buildPath, 'Cmake', 'source.cmake')

            if not os.path.exists(source_cmake_path):
                Util.reportError(f"source.cmake not found at: {source_cmake_path}")

            try:
                with open(source_cmake_path, 'r', encoding='utf-8') as f:
                    content = f.read()

                import re

                pattern = r'set\(source(.*?)\)'
                match = re.search(pattern, content, re.S | re.M)

                if not match:
                    patterns = [
                        r'set\(\s*source(.*?)\)',
                        r'set\(SOURCES?(.*?)\)',
                        r'set\(\s*SOURCES?(.*?)\)',
                    ]

                    for pattern in patterns:
                        match = re.search(pattern, content, re.S | re.M)
                        if match:
                            break

                if not match:
                    Util.reportError("Could not find set(source ... ) in source.cmake")

                source_content = match.group(1)
                raw_lines = source_content.split('\n')
                data = [i.replace('\\', '/') for i in raw_lines if i.replace(' ', '') != '']

            except Exception as e:
                return "\n\t"

        elif (compilerEnv == 'emake'):
            build_src_path = os.path.join(self.buildPath, 'cfg', 'file', 'build_src.txt')

            if not os.path.exists(build_src_path):
                Util.reportError(f"build_src.txt not found at: {build_src_path}")

            import re
            with open(build_src_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()
                for line in lines:
                    line = line.strip()
                    if line:
                        # Extract file path, remove |1 and other flags
                        file_path = re.compile(r'^(.*?)(?:\|1)?$').sub(r'\1', line).replace('\\', '/')
                        # Ensure path starts with ./ (this is the format Polyspace expects)
                        if not file_path.startswith('./'):
                            file_path = './' + file_path
                        data.append(file_path)

        return "\n\t".join(data)

    def getConfig(self):
        tem = '<element>{}</element>'
        config = [str(i).strip() for i in self.allConfig['polyspace']['option'].split("|")]
        all = []
        for i in config:
            all.append(tem.format(i))
        return '\n\t\t'.join(all)


if __name__ == '__main__':
    args = sys.argv[1:]
    workModeParm = 'normal'
    analysisModeParm = 'specify'
    changeFilePathParm = ''

    # get args,e.g -mode=ci -analysismode=full -file='./changedFiles.txt'
    for arg in args:
        if arg.startswith('-mode='):
            workModeParm = arg.split('=')[1]
        elif arg.startswith('-analysismode='):
            analysisModeParm = arg.split('=')[1]
        elif arg.startswith('-file='):
            changeFilePathParm = arg.split('=')[1]
            possible_paths = [
                os.path.join(Util.getBuildPath(), changeFilePathParm),
                os.path.join(Util.getProjectPath(), 'Build', changeFilePathParm),
                os.path.join(Util.getProjectPath(), 'Customer', 'Build', changeFilePathParm),
                changeFilePathParm  # Absolute path
            ]

            for path in possible_paths:
                if os.path.isfile(path):
                    polyspacechangeFilePath = path
                    break
            else:
                Util.reportError(
                    f'-file={changeFilePathParm} not exist. Please check the file and confirm if the input format is -file=\'./changedFiles.txt\'.')

    if workModeParm == 'ci':
        polyspaceWorkMode = 'ci'
    elif workModeParm == 'normal':
        polyspaceWorkMode = 'normal'
    else:
        Util.reportError(f'-mode={workModeParm} not support,Only Support -mode=noraml,-mode=ci')
        raise ValueError

    if analysisModeParm == 'full':
        polyspaceAnalysisMode = 'full'
    elif analysisModeParm == 'specify':
        polyspaceAnalysisMode = 'specify'
    elif analysisModeParm == 'guard':
        polyspaceAnalysisMode = 'guard'
    else:
        Util.reportError(f'-analysismode={analysisModeParm} not support ')
        raise ValueError

    Util.reportNormal(f'Current working mode: {polyspaceWorkMode}')
    Util.reportNormal(f'Current analysis mode: {polyspaceAnalysisMode}')
    Util.reportNormal(f'Current change file: {polyspacechangeFilePath}')

    # compiler select
    possible_build_bat_paths = [
        os.path.join(Util.getBuildPath(), 'Tools', 'Build.bat'),
        os.path.join(Util.getBuildPath(), 'tools', 'Build.bat'),
        os.path.join(Util.getProjectPath(), 'Build', 'Tools', 'Build.bat'),
        os.path.join(Util.getProjectPath(), 'Customer', 'Build', 'Tools', 'Build.bat'),
    ]

    buildBatPath = None
    for path in possible_build_bat_paths:
        if os.path.exists(path):
            buildBatPath = path
            break

    if not buildBatPath:
        Util.reportError("Build.bat not found in any location")

    fd = open(buildBatPath, 'r', encoding='utf-8')
    text = fd.read()
    fd.close()

    if (re.compile(r'Emake\.exe', re.IGNORECASE).findall(text) != []):
        compilerEnv = 'emake'
        Util.reportNormal(f"Found 'Emake.exe' in {buildBatPath}")
    else:
        # Check for CMake indicators
        if re.compile(r'cmake|CMake', re.IGNORECASE).findall(text):
            compilerEnv = 'cmake'

    Util.reportNormal(f"Compiler select: {compilerEnv}")

    ps = PolyspaceUtil()
    ps.start()