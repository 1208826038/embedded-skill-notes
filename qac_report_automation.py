import re
import sys
import queue
import shutil
import signal
import psutil
import os.path
import configparser
import subprocess
import threading
from colorama import Fore, init
from xml.dom.minidom import parse, parseString
from lxml import html, etree
import xml.etree.ElementTree as ET
import datetime
import getpass
from openpyxl import load_workbook
sys.path.insert(0, os.path.join(os.path.dirname(os.path.realpath(__file__)), "schema"))
from Reports_pb2 import QAReportsData
from ReportHelpers import *
from google.protobuf import text_format

G = Global()
G.data = QAReportsData()

_LinkPath = None

sys.path.insert(0,os.getcwd())
workMode = 'normal'
analysisMode = 'specify'
changeFile = ''
suffixesSource = ['_Cfg.c', '_Lcfg.c', '_Cbk.c', '_PBcfg.c', '_define.c', '_Callout.c', '_Callout_Stubs.c', '_Hal_Core.c', '_Irq.c', '.c']
suffixesInclude = ['_Cfg.h', '_Lcfg.h', '_Cbk.h', '_PBcfg.h', '_define.h', '_Callout.h', '_Callout_Stubs.h', '_Hal_Core.h', '_GeneralTypes.h', '_Type.h', '_Types.h', '_Memmap.h', '_Int.h', '_Irq.h', '.h']
prefixesKey = ['SchM_', 'Rte_']
sourceKeyWords = []

# compiler select, e.g. cmake, emake
compilerEnv = 'cmake'

init(autoreset = True)

class Util:

    def __init__(self):
        pass

    @staticmethod
    def greenColor():
        return ""#"\033[1;32;0m  "

    @staticmethod
    def redColor():
        return ""#"\033[1;31;0m  "

    @staticmethod
    def getProjectPath():
        """获取项目根目录（兼容两种结构）"""
        current_dir = sys.path[0]

        # 1. 检查是否是嵌套结构（Customer/Build）
        if "Customer\\Build" in current_dir or "Customer/Build" in current_dir:
            # 嵌套结构：向上两级到项目根目录
            project_root = os.path.dirname(os.path.dirname(current_dir))
            return project_root

        # 2. 检查是否是SourceCode结构（Build目录的父目录包含SourceCode）
        if "\\Build" in current_dir:
            parent_dir = current_dir.split("\Build")[0]
            sourcecode_path = os.path.join(parent_dir, "SourceCode")
            if os.path.exists(sourcecode_path):
                # SourceCode结构
                return parent_dir

        # 3. 如果上述都不符合，使用原始逻辑
        return sys.path[0].split("\Build")[0] if "\Build" in sys.path[0] else sys.path[0]

    @staticmethod
    def getSourceRelPath():
        """获取源代码相对路径"""
        if compilerEnv == 'cmake':
            if os.path.exists(os.path.join('..', 'SourceCode', 'CMakeLists.txt')):
                relPath = '/SourceCode/'
            else:
                relPath = '/../'
        elif compilerEnv == 'emake':
            relPath = '/'
        else:
            relPath = '/SourceCode/'
        return relPath

    @staticmethod
    def getBuildPath():
        """获取Build目录路径（兼容两种结构）"""
        current_dir = sys.path[0]

        # 如果当前目录包含"Build"，直接使用
        if "Build" in os.path.basename(current_dir):
            return current_dir

        # 否则查找Build目录
        project_root = Util.getProjectPath()

        # 尝试不同位置的Build目录
        possible_build_paths = [
            os.path.join(project_root, "Build"),  # 原始结构
            os.path.join(project_root, "Customer", "Build"),  # 嵌套结构
            current_dir  # 回退
        ]

        for path in possible_build_paths:
            if os.path.exists(path):
                print(f"DEBUG: Using Build path: {path}")
                return path

        # 如果都找不到，使用当前目录
        print(f"DEBUG: Build path not found, using current dir: {current_dir}")
        return current_dir

    @staticmethod
    def getQacEnvSource():
        return Util.getBuildPath() + r'/Tools/Components/QAC/Project'

    @staticmethod
    def getQacEnvDestination():
        return Util.getBuildPath() + r'/CodeVerify/CodeStaticCheck'

    @staticmethod
    def getQacTemplatePath():
        return Util.getBuildPath() + r'/Tools/Components/QAC/templates/prqaproject/prqaproject_template.xml'

    @staticmethod
    def getQacReportTemplatePath():
        return Util.getBuildPath() + r'/Tools/Components/QAC/templates/report/report_template.html'

    @staticmethod
    def getConfigPath():
        """获取config.ini文件路径"""
        # 优先在Build目录下查找
        build_path = Util.getBuildPath()
        config_path = os.path.join(build_path, "VerifyCfg", "config.ini")

        if os.path.exists(config_path):
            print(f"DEBUG: Found config.ini at: {config_path}")
            return config_path

        # 如果Build目录下没有，尝试项目根目录下的Build
        project_root = Util.getProjectPath()
        alternative_paths = [
            os.path.join(project_root, "Build", "VerifyCfg", "config.ini"),
            os.path.join(project_root, "Customer", "Build", "VerifyCfg", "config.ini"),
            os.path.join(project_root, "VerifyCfg", "config.ini"),
        ]

        for path in alternative_paths:
            if os.path.exists(path):
                print(f"DEBUG: Found config.ini at: {path}")
                return path

        # 如果都找不到，返回原始路径
        print(f"DEBUG: Config.ini not found, returning default: {config_path}")
        return config_path

    @staticmethod
    def getCCTPath():
        return Util.getBuildPath() + r'/CodeVerify/CodeStaticCheck/prqa/configs/Initial/config/cct/'

    @staticmethod
    def getCCTTargetPath():
        return Util.getBuildPath() + r'/CodeVerify/CodeStaticCheck/prqa/configs/Initial/config/'

    @staticmethod
    def getSourceRelPath():
        if compilerEnv == 'cmake':
            if os.path.exists(os.path.join('..', 'SourceCode', 'CMakeLists.txt')):
                relPath = '/SourceCode/'
            else:
                relPath = '/../'
        elif compilerEnv == 'emake':
            relPath = '/'
        else:
            relPath = '/SourceCode/'
        return relPath

    @staticmethod
    def getSourcePath():
        """获取源代码基础路径（兼容两种结构）"""
        project_root = Util.getProjectPath()

        # 检查项目结构类型
        # 先检查是否是SourceCode结构
        sourcecode_path = os.path.join(project_root, "SourceCode")
        if os.path.exists(sourcecode_path):
            # SourceCode结构：源代码在SourceCode目录下
            return sourcecode_path.replace("\\", "/") + "/"

        # 如果不是SourceCode结构，则是嵌套结构
        # 嵌套结构：源代码直接在项目根目录下（BSW、Customer等）
        return project_root.replace("\\", "/") + "/"

    @staticmethod
    def resolveRelativePath(rel_path):
        """
        解析相对路径（兼容两种结构）
        rel_path: 如 "./BSW/ManualCode/Cdd/FaultM/FaultM.c"
        返回: 绝对路径
        """
        if not rel_path.startswith('./'):
            return rel_path

        # 获取源代码基础路径
        source_path = Util.getSourcePath()
        relative_part = rel_path[2:]  # 去掉"./"

        # 构建绝对路径
        abs_path = os.path.join(source_path, relative_part).replace("\\", "/")

        return abs_path
    @staticmethod
    def getQacCheckFilePath():
        """获取qac_check.ini文件路径"""
        build_path = Util.getBuildPath()
        qac_check_path = os.path.join(build_path, "VerifyCfg", "qac_check.ini")

        if os.path.exists(qac_check_path):
            print(f"DEBUG: Found qac_check.ini at: {qac_check_path}")
            return qac_check_path

        # 尝试其他可能位置
        project_root = Util.getProjectPath()
        alternative_paths = [
            os.path.join(build_path, "QacCheck", "qac_check.ini"),  # 您提到的目录
            os.path.join(build_path, "qac_check.ini"),
            os.path.join(project_root, "Build", "VerifyCfg", "qac_check.ini"),
            os.path.join(project_root, "Customer", "Build", "QacCheck", "qac_check.ini"),
        ]

        for path in alternative_paths:
            if os.path.exists(path):
                print(f"DEBUG: Found qac_check.ini at: {path}")
                return path

        print(f"DEBUG: qac_check.ini not found, returning default: {qac_check_path}")
        return qac_check_path

    @staticmethod
    def getTemAcfPath():
        return Util.getBuildPath() + r'/Tools/Components/QAC/templates/acf/acf_template.acf'

    @staticmethod
    def getTargetAcf():
        return Util.getBuildPath() + r'/CodeVerify/CodeStaticCheck/prqa/configs/Initial/config/catl.acf'

    @staticmethod
    def getTargetPRQA():
        return Util.getBuildPath() + r'/CodeVerify/CodeStaticCheck/prqaproject.xml'

    @staticmethod
    def writeXmlFile(filePath, document):
        with open(filePath, 'w', encoding='utf-8') as f:
            document.writexml(f, indent=' ', newl='\n', addindent=' ', encoding='utf-8')

    @staticmethod
    def getCipPath():
        return Util.getBuildPath() + '/CodeVerify/CodeStaticCheck/prqa/configs/initial/cip'

    @staticmethod
    def reportError(inputStr):
        print(Fore.RED + inputStr)
        exit(-1)

    @staticmethod
    def reportErrorNoExit(inputStr):
        print(Fore.RED + inputStr)

    @staticmethod
    def reportNormal(inputStr):
        print(Fore.GREEN + inputStr)


class sType:
    COMMENT_SUPPRESSION = 0
    PRAGMA_SUPPRESSION = 1
    BASELINE_SUPPRESSION = 2
    MACRO_SUPPRESSION = 3
    INTERACTIVE_SUPPRESSION = 4

class OptionalColumns:
    def __init__(self, suppressions):
        self.has_comment = False
        self.has_macro = False
        for s in suppressions:
            if s.macroName:
                self.has_macro = True
            if s.comment:
                self.has_comment = True

    def write_row(self, suppression):
        ret = ''
        if suppression.macroName:
            ret += '<td>%s</td>' % suppression.macroName
        else:
            if self.has_macro:
                ret += '<td></td>'

        if suppression.comment:
            ret += '<td>%s</td>' % suppression.comment
        else:
            if self.has_comment:
                ret += '<td></td>'
        return ret

    def write_header(self):
        ret = ''
        if self.has_macro:
            ret += '<th>Macro name</th>'
        if self.has_comment:
            ret += '<th>Justification comment</th>'
        return ret

class QacUtil:

    def __init__(self):
        self.reportList = ['CRR', 'HMR', 'MCR', 'MDR', 'RCR','SCR', 'SUR']
        self.generalConfig = configparser.ConfigParser()

        if(not os.path.exists(Util.getConfigPath())):
            Util.reportError(f'ERROR: Can\'t find file config.ini')
            raise ValueError

        self.generalConfig.read(Util.getConfigPath())

        # Create Project environment
        if (not os.path.exists(Util.getQacEnvSource())):
            Util.reportError("ProjectEnv source don't exits, path {path}".format(path = Util.getQacEnvSource()))
            raise ValueError

        if (not os.path.exists(Util.getQacEnvDestination())):
            os.makedirs(Util.getQacEnvDestination())

        shutil.copytree(Util.getQacEnvSource(), Util.getQacEnvDestination(), dirs_exist_ok=True)

        # Check cct
        if(not self.generalConfig.has_option('qac', 'cct')):
            Util.reportError(f'ERROR: Can\'t find cct in file config.ini')
            raise ValueError
        elif(not os.path.exists(Util.getCCTPath() + '/' + self.generalConfig.get("qac", "cct"))):
            Util.reportError(f'ERROR: Can\'t find cct file in path {Util.getCCTPath()}')
            raise ValueError


        # Check options
        if(not self.generalConfig.has_option('qac', 'option')):
            Util.reportError(f'ERROR: Can\'t find option in file config.ini')
            raise ValueError
        else:
            marcos = self.generalConfig.get('qac', 'option').split('|')
            if(len(marcos) == 0):
                Util.reportError(f'ERROR: Can\'t find option in file config.ini')
                raise ValueError

        # Get QAC config file
        self.QACRules = Util.getBuildPath() + "\\Tools\\Components\\QAC\\doc\\QAC_Rule.xlsx"

    def start(self):
        print(f"脚本运行目录: {sys.path[0]}")
        print(f"项目根目录: {Util.getProjectPath()}")
        print(f"Build目录: {Util.getBuildPath()}")
        print(f"源代码目录: {Util.getSourcePath()}")
        print(f"QAC配置文件: {Util.getQacCheckFilePath()}")
        self.qacPathCheck()

        # version check QAC 2023
        self.versionCheck()

        self.dom = self.readXML(Util.getQacTemplatePath())

        self.setRootPaths()

        self.setCCT()

        self.getCfgEnv()

        self.setInclude()

        ret = self.setCheckFile()

        self.setMarcos()

        self.writeXML()

        # delete cip file
        self.delCipFile()

        if ret == 0:
            # execute QAC command
            self.qacAnaze()
            # delete cip file
            self.delCipFile()
            # report analysis report
            self.analysisReport()
            Util.reportNormal('<SUCCESS>: CATL Qac Analysys success!')
        else:
            Util.reportNormal('<INFO>: The specified project does not have any source files!Please check the project')

    def getCfgEnv(self):
        data = []
        try:
            if (compilerEnv == 'cmake'):
                with open(os.path.join(Util.getBuildPath(), 'Cmake/env.cmake'), 'r', encoding='utf-8') as f:
                    data = f.read()
                # Get Compiler     greenhills_arm
                self.buildTool = re.compile('set\(COMPILER (.*?)\)').findall(data)[0]
                # Get CHIP
                self.chipTool = re.compile('set\(CHIP_NAME (.*?)\)').findall(data)[0]
                # Get projectName
                self.projectName = re.compile('set\(PROJECT_NAME (.*?)\)').findall(data)[0]
                # Get compiler Path
                if(os.environ.get("CI_COMPILER_BASE", default=None) is not None):
                    self.compilerBase = os.environ.get("CI_COMPILER_BASE", default=None).replace("\\", "/")
                    Util.reportNormal("Compiler: "+ self.buildTool + " CHIP: "+ self.chipTool +
                        "\nCI_COMPILER_BASE: "+self.compilerBase + " PROJECT_NAME: " + self.projectName)
                else:
                    self.compilerBase = re.compile('set\(COMPILER_BASE "(.*?)"\)').findall(data)[0].replace("\\", "/")
                    Util.reportNormal("Compiler: "+ self.buildTool + "  CHIP: "+ self.chipTool +
                        "\nCOMPILER_BASE: " + self.compilerBase + "  PROJECT_NAME: " + self.projectName)

            elif (compilerEnv == 'emake'):
                with open(os.path.join(Util.getBuildPath(), 'cfg/project.ini'), 'r', encoding='utf-8') as f:
                    data = f.read()
                # Get Compiler
                self.buildTool = re.compile('build_toolchain\s*=\s*(.*)').findall(data)[0]
                # Get CHIP
                self.chipTool = re.compile('chip\s*=\s*(.*)').findall(data)[0]
                # Get projectName
                self.projectName = re.compile('project_name\s*=\s*(.*)').findall(data)[0]
                # Get compiler Path
                if(os.environ.get("CI_COMPILER_BASE", default=None) is not None):
                    self.compilerBase = os.environ.get("CI_COMPILER_BASE", default=None).replace("\\", "/")
                    Util.reportNormal("build_toolchain: "+ self.buildTool + "  chip: "+ self.chipTool + \
                        "\nCI_COMPILER_BASE: " + self.compilerBase + "  project_name: " + self.projectName)
                else:
                    self.compilerBase = re.compile('compiler_path\s*=\s*(.*)').findall(data)[0].replace("\\", "/")
                    Util.reportNormal("build_toolchain: "+ self.buildTool + "  chip: "+ self.chipTool + \
                        "\nCOMPILER_BASE: " + self.compilerBase + "  project_name: " + self.projectName)
        except FileNotFoundError:
            StackAnalysisCommmonUtil.reportError(f"compilerEnv: + {compilerEnv}, env.cmake or cfg/project.ini not found")
        except Exception as e:
            StackAnalysisCommmonUtil.reportError(f"Error reading cfg env: {str(e)}, {type(e).__name__} at line {sys.exc_info()[-1].tb_lineno}")

    def getCompilerPath(self):
        retInc = []

        if (compilerEnv == 'cmake'):
            if (not os.path.exists(Util.getBuildPath() + f'/Tools/Components/toolchains/{self.buildTool}.cmake')):
                Util.reportError(f'Can\'t find {self.buildTool}.cmake in Toolchains')
            with open(Util.getBuildPath() + f'/Tools/Components/toolchains/{self.buildTool}.cmake', 'r', encoding='utf-8') as f:
                raw = re.compile('set\(COMPILER_INC  "(.*?)"\)', re.S).findall(f.read())[0].replace(' ', '')
                allTem = raw.split('\n')

                for index in allTem:
                    retInc.append(index.replace(r'${COMPILER_BASE}',self.compilerBase).strip() + "/")
        elif (compilerEnv == 'emake'):
            compilerPath = self.compilerBase.strip() + "/ctc/include" + "/"
            compilerPath = compilerPath.replace("/ctc/ctc", "/ctc")
            retInc.append(compilerPath)

        return retInc

    def readXML(self, filePath):
        with open(filePath, 'r', encoding='utf-8') as f:
            text = f.read()
            text = re.sub('\n\\s+', '', text)
            doc = parseString(text)
            return doc

    def setRootPaths(self):
        rootPaths = self.dom.getElementsByTagName('root_paths')[0]
        while rootPaths.childNodes != []:
            rootPaths.removeChild(rootPaths.childNodes[0])

        sourceRoot = self.dom.createElement('root_path')
        sourceRoot.setAttribute('path', Util.getProjectPath())
        sourceRoot.setAttribute('name', 'SOURCE_ROOT')
        rootPaths.appendChild(sourceRoot)

    def setCCT(self):
        ccts = self.dom.getElementsByTagName('ccts')[0]
        while ccts.childNodes != []:
            ccts.removeChild(ccts.childNodes[0])

        cctFile = self.generalConfig.get('qac', 'cct')  # may verify cct file exits first?
        if(not os.path.exists(os.path.join(Util.getCCTPath(), cctFile))):
            Util.reportError("CCT File don't exits, int path {path}".format(path = Util.getCCTPath()))
            raise ValueError

        cct = self.dom.createElement('cct')
        cct.setAttribute('target', 'C')
        cct.setAttribute('active', 'yes')
        cct.setAttribute('name', cctFile)
        ccts.appendChild(cct)

        src = Util.getCCTPath() + cctFile
        dst = Util.getCCTTargetPath()
        shutil.copy(src, dst)

    def setInclude(self):
        """
        设置include目录（兼容两种结构）
        """
        includes = self.dom.getElementsByTagName('includes')[0]
        while includes.childNodes:
            includes.removeChild(includes.childNodes[0])

        includeFileNum = 0
        lines = self.readInclude().split('\n\t')

        for line in lines:
            line = line.strip().replace("\\", "/")
            if not line:
                continue

            # 解析include路径
            if line.startswith('./'):
                includePath = self.resolveQacFilePath(line)
            else:
                includePath = line

            if not os.path.exists(includePath):
                if compilerEnv != 'emake':
                    # 对于cmake环境，尝试查找可能的替代路径
                    continue

            includeEle = self.dom.createElement('include')
            includeEle.setAttribute('path', includePath)
            includes.appendChild(includeEle)
            includeFileNum += 1

        # 添加编译器路径
        compilerList = self.getCompilerPath()
        for compiler in compilerList:
            compiler = compiler.replace("\\", "/").strip()
            compilerPath = self.correct_case_path(compiler)
            if not os.path.exists(compilerPath):
                continue

            includeEle = self.dom.createElement('include')
            includeEle.setAttribute('path', compilerPath)
            includes.appendChild(includeEle)
            includeFileNum += 1

        if includeFileNum == 0:
            Util.reportError('ERROR: No include paths found!')
            raise ValueError

    # read all include to the project file
    def readInclude(self):
        print(Fore.GREEN + "读取Include文件")
        data = []
        if (compilerEnv == 'cmake'):
            include_cmake = os.path.join(Util.getBuildPath(), 'Cmake/include.cmake')
            print(f"读取include.cmake: {include_cmake}")
            if os.path.exists(include_cmake):
                with open(include_cmake, 'r', encoding='utf-8') as f:
                    content = f.read()
                    matches = re.compile(r'set\(include_dir(.*?)\)', re.S | re.M).findall(content)
                    if matches:
                        raw_lines = matches[0].split('\n')
                        for line in raw_lines:
                            line = line.strip().replace('\\', '/')
                            if line and not line.startswith('#'):
                                data.append(line)
            else:
                print(Fore.YELLOW + f"警告: {include_cmake} 不存在")
        elif (compilerEnv == 'emake'):
            include_dir_txt = os.path.join(Util.getBuildPath(), 'cfg/file/include_dir.txt')
            print(f"读取include_dir.txt: {include_dir_txt}")
            if os.path.exists(include_dir_txt):
                with open(include_dir_txt, 'r', encoding='utf-8') as f:
                    for line in f.readlines():
                        line = line.strip().replace('\\', '/')
                        if line:
                            data.append(line)
            else:
                print(Fore.YELLOW + f"警告: {include_dir_txt} 不存在")

        print(f"找到 {len(data)} 个include路径")
        return "\n\t".join(data)

    # read all source to the project file
    def readSource(self):
        print(Fore.GREEN + "读取Source文件")
        data = []
        if (compilerEnv == 'cmake'):
            source_cmake = os.path.join(Util.getBuildPath(), 'Cmake/source.cmake')
            print(f"读取source.cmake: {source_cmake}")
            if os.path.exists(source_cmake):
                with open(source_cmake, 'r', encoding='utf-8') as f:
                    content = f.read()
                    matches = re.compile(r'set\(source(.*?)\)', re.S | re.M).findall(content)
                    if matches:
                        raw_lines = matches[0].split('\n')
                        for line in raw_lines:
                            line = line.strip().replace('\\', '/')
                            if line and not line.startswith('#'):
                                data.append(line)
            else:
                print(Fore.YELLOW + f"警告: {source_cmake} 不存在")
        elif (compilerEnv == 'emake'):
            build_src_txt = os.path.join(Util.getBuildPath(), 'cfg/file/build_src.txt')
            print(f"读取build_src.txt: {build_src_txt}")
            if os.path.exists(build_src_txt):
                with open(build_src_txt, 'r', encoding='utf-8') as f:
                    for line in f.readlines():
                        line = line.strip()
                        if line:
                            # 移除可能存在的 "|1" 后缀
                            line = re.sub(r'^(.*?)(?:\|1)?$', r'\1', line).replace('\\', '/')
                            data.append(line)
            else:
                print(Fore.YELLOW + f"警告: {build_src_txt} 不存在")

        print(f"找到 {len(data)} 个源文件")
        return "\n\t".join(data)

    def qacPathCheck(self):
        if(os.environ.get("CI_HELIXQAC_BASE", default=None) is not None):
            self.helixQacCmd = os.environ.get("CI_HELIXQAC_BASE", default=None) + r'/common/lib/qacli.exe'
        elif(self.generalConfig.get('qac', 'helix_qac') != None):
            self.helixQacCmd = self.generalConfig.get('qac', 'helix_qac') + r'/common/lib/qacli.exe'
        else:
            Util.reportError('Helix Qac Path do not config!!!')
            raise ValueError
        if(not os.path.exists(self.helixQacCmd)):
            Util.reportError(f'Helix Qac Path:{self.helixQacCmd} do not exists!!!')
            raise ValueError

    def versionCheck(self):
        #version check
        checkCmd = self.helixQacCmd + ' --version'
        qacVersion = subprocess.run(checkCmd, capture_output = True).stdout.decode('utf-8')

        if ((qacVersion.find("2023") != -1) or (qacVersion.find("2025") != -1)):
            Util.reportNormal(f'Current QAC Version: {qacVersion}')
        else:
            Util.reportError(
                'Your QAC is not the version of 2023.2/2025.2. Install it from the link: http://bms-wiki.catlbattery.com:81/UserManual_QAC')
            raise ValueError

    def correct_case_path(self, original_path):
        """修正路径中的大小写错误"""
        parts = original_path.split('/')
        corrected_parts = []
        current_path = ""

        for part in parts:
            if not part:  # 跳过空部分
                continue
            current_path = os.path.join(current_path, part) if current_path else part

            # 尝试修正大小写
            parent = os.path.dirname(current_path) if current_path != part else ""

            try:
                candidates = [f for f in os.listdir(parent) if f.lower() == part.lower()]
            except FileNotFoundError:
                candidates = []

            if candidates:
                corrected_part = candidates[0]  # 取第一个匹配项
                corrected_parts.append(corrected_part)
                current_path = os.path.join(parent, corrected_part) if parent else corrected_part
            else:
                if ':' in part:
                    current_path = current_path + '/'
                corrected_parts.append(part)  # 没有匹配项，保持原样

        return '/'.join(corrected_parts)

    def setCheckFile(self):
        """
        设置要分析的文件
        """
        files = self.dom.getElementsByTagName('files')[0]
        while files.childNodes:
            files.removeChild(files.childNodes[0])

        global sourceKeyWords
        lineKey = False
        suppressPath = []
        defaultKey = []
        fullAnalysisPath = []

        # 读取FullAnalysis部分
        with open(Util.getQacCheckFilePath(), 'r', encoding='utf-8') as f:
            lines = f.readlines()
            for line in lines:
                line = line.replace("\\", "/").strip()
                if not lineKey:
                    if not line.startswith(r"["):
                        continue
                    if line.startswith(r"[FullAnalysis]"):
                        lineKey = True
                elif lineKey and line.startswith(r"["):
                    break
                else:
                    if not line.startswith('./'):
                        continue
                    fullAnalysisPath.append(line)

        # 读取项目源文件
        projectSource = self.readSource().split('\n\t')
        projectInclude = self.readInclude().split('\n\t')

        sourceKey = []
        suffixesSourceLower = [s.lower() for s in suffixesSource]
        suffixesIncludeLower = [s.lower() for s in suffixesInclude]
        prefixesKeyLower = [s.lower() for s in prefixesKey]

        lineKey = False
        with open(Util.getQacCheckFilePath(), 'r', encoding='utf-8') as f:
            lines = f.readlines()
            for line in lines:
                line = line.replace("\\", "/").strip()
                if not lineKey:
                    if not line.startswith(r"["):
                        continue
                    if line.startswith(r"[SpecifyAnalysis]") and analysisMode == "specify":
                        lineKey = True
                    if line.startswith(r"[SuppressPath]") and analysisMode == "full":
                        lineKey = True
                    if line.startswith(r'[SuppressPath]') and analysisMode == "guard":
                        lineKey = True
                elif lineKey and line.startswith(r"["):
                    break
                else:
                    if not line.startswith('./'):
                        continue

                    if (analysisMode == "guard") or (analysisMode == 'full'):
                        suppressPath.append(line)
                    else:
                        # 使用统一的路径解析方法
                        filePath = self.resolveQacFilePath(line)

                        if not os.path.exists(filePath):
                            Util.reportError(f'Specify file not exists! {filePath}')
                            raise ValueError

                        # 获取文件所在目录
                        file_dir = os.path.dirname(filePath)

                        fileEle = self.dom.createElement('file')
                        fileEle.setAttribute('target', 'C')
                        fileEle.setAttribute('name', os.path.basename(filePath))
                        fileEle.setAttribute('folder', file_dir.replace('\\', '/'))
                        files.appendChild(fileEle)

                        # 提取源文件关键字
                        filename_lower = os.path.basename(filePath).lower()
                        for suffix in suffixesSourceLower:
                            if filename_lower.endswith(suffix):
                                sourceItem = filename_lower[:-len(suffix)]

                                # 过滤前缀
                                for prefix in prefixesKeyLower:
                                    if sourceItem.startswith(prefix):
                                        sourceItem = sourceItem[len(prefix):]
                                        break

                                # 添加源关键字
                                if sourceItem not in sourceKey:
                                    sourceKey.append(sourceItem)
                                break

        # 处理guard和full分析模式
        if (analysisMode == "guard") or (analysisMode == 'full'):
            # 处理每个源文件
            for sourceLine in projectSource:
                sourceLine = sourceLine.replace("\\", "/").strip()
                if not sourceLine.startswith('./'):
                    continue

                # 解析源文件路径
                sourceFilePath = self.resolveQacFilePath(sourceLine)

                # 检查是否在抑制路径中
                suppressEnable = False
                for suppress in suppressPath:
                    suppressFilePath = self.resolveQacFilePath(suppress)
                    if suppressFilePath in sourceFilePath:
                        suppressEnable = True
                        break

                if not suppressEnable:
                    # 检查是否在完整分析路径中
                    for fullAnalysis in fullAnalysisPath:
                        fullAnalysisPathResolved = self.resolveQacFilePath(fullAnalysis)

                        if analysisMode == "guard":
                            # 检查文件是否匹配关键字
                            filename_lower = os.path.basename(sourceFilePath).lower()
                            key_found = False

                            for key in sourceKey:
                                key_lower = key.lower()
                                for suffix in suffixesSourceLower:
                                    if (key_lower + suffix) == filename_lower:
                                        key_found = True
                                        break
                                    # 检查带前缀的
                                    for prefix in prefixesKeyLower:
                                        if (prefix + key_lower + suffix) == filename_lower:
                                            key_found = True
                                            break
                                if key_found:
                                    break

                            if key_found and fullAnalysisPathResolved in sourceFilePath:
                                if os.path.exists(sourceFilePath):
                                    file_dir = os.path.dirname(sourceFilePath)
                                    fileEle = self.dom.createElement('file')
                                    fileEle.setAttribute('target', 'C')
                                    fileEle.setAttribute('name', os.path.basename(sourceFilePath))
                                    fileEle.setAttribute('folder', file_dir.replace('\\', '/'))
                                    files.appendChild(fileEle)
                                break

                        elif analysisMode == 'full':
                            if fullAnalysisPathResolved in sourceFilePath:
                                if os.path.exists(sourceFilePath):
                                    file_dir = os.path.dirname(sourceFilePath)
                                    fileEle = self.dom.createElement('file')
                                    fileEle.setAttribute('target', 'C')
                                    fileEle.setAttribute('name', os.path.basename(sourceFilePath))
                                    fileEle.setAttribute('folder', file_dir.replace('\\', '/'))
                                    files.appendChild(fileEle)

                                    # 提取源文件关键字
                                    filename_lower = os.path.basename(sourceFilePath).lower()
                                    for suffix in suffixesSourceLower:
                                        if filename_lower.endswith(suffix):
                                            sourceItem = filename_lower[:-len(suffix)]
                                            for prefix in prefixesKeyLower:
                                                if sourceItem.startswith(prefix):
                                                    sourceItem = sourceItem[len(prefix):]
                                                    break
                                            if sourceItem not in sourceKey:
                                                sourceKey.append(sourceItem)
                                            break
                                break

        sourceKeyWords = sourceKey

        # 检查是否添加了文件
        files_count = len(self.dom.getElementsByTagName('files')[0].childNodes)

        if files_count == 0:
            return -1
        else:
            return 0

    def setMarcos(self):
        self.acfDoc = self.readXML(Util.getTemAcfPath())
        inputTo = self.acfDoc.getElementsByTagName('input_to')[0]
        marcos = self.generalConfig.get('qac', 'option').split('|')
        # Add marco
        for marco in marcos:
            optionEle = self.acfDoc.createElement('option')
            optionEle.setAttribute('name', '-d ')
            optionEle.setAttribute('argument', marco)
            inputTo.appendChild(optionEle)

        # Suppress File Check
        suppressPath = []

        # Add Suppress analysis file
        lineKey = False
        with open(Util.getQacCheckFilePath(), 'r', encoding='utf-8')  as f:
            for line in f.readlines():
                line = line.replace("\\", "/")
                if not lineKey:
                    if line.startswith('[SuppressPath]'):
                        lineKey = True
                        continue
                elif  lineKey == True and line.startswith(r'['):
                    break
                else:
                    if not line.startswith('./'):
                        continue
                    suppressPath.append(line)

        for suppress in suppressPath:
            suppressDir = Util.getSourcePath() + suppress.replace("\n",'')
            if os.path.isdir(suppressDir):
                optionEle = self.acfDoc.createElement('option')
                optionEle.setAttribute('name', '-quiet ')
                optionEle.setAttribute('argument', suppress.replace("\n",'').replace('./', '${SOURCE_ROOT}' + Util.getSourceRelPath()))
                inputTo.appendChild(optionEle)

    def writeXML(self):
        Util.writeXmlFile(Util.getTargetPRQA(), self.dom)
        Util.writeXmlFile(Util.getTargetAcf(), self.acfDoc)

    def delCipFile(self):
        cipPath = Util.getCipPath()
        if(not os.path.exists(cipPath)):
            return
        for file in os.listdir(cipPath):
            if(os.path.exists(file)):
                os.remove(cipPath+'/'+file)

    def qacAnaze(self):
        Util.reportNormal('Start QAC analysis')
        #analyze
        anazeCmd = self.helixQacCmd + ' analyze -cf -P '+ Util.getBuildPath()+ '/CodeVerify/CodeStaticCheck'
        Util.reportNormal(anazeCmd)
        anazeRet = subprocess.run(anazeCmd, shell=True,text=True)
        if anazeRet.returncode != 0:
            Util.reportError("QAC  Analysis ERROR !!!")
            raise ValueError

        #generate report
        reprotRet = Util.reportNormal('QAC  Analysis Success, Start QAC report generation')
        reportCmd = self.helixQacCmd + ' report -P ' + Util.getBuildPath() + '/CodeVerify/CodeStaticCheck --type '
        Util.reportNormal(reportCmd)
        for report in self.reportList:
            reportRet = subprocess.run(reportCmd + report, shell=True)
            if reportRet.returncode != 0:
                Util.reportError("QAC  Report ERROR !!!")

    def resolveQacFilePath(self, rel_path):
        """
        解析qac_check.ini中的文件路径（兼容两种结构）
        rel_path: 如 "./BSW/ManualCode/Cdd/FaultM/FaultM.c"
        返回: 绝对路径
        """
        rel_path = rel_path.strip().replace("\\", "/")
        if not rel_path.startswith('./'):
            return rel_path

        # 使用Util的路径解析方法
        abs_path = Util.resolveRelativePath(rel_path)

        # 修正大小写
        corrected_path = self.correct_case_path(abs_path)

        return corrected_path
    def get_git_remote_url(remote_name: str = "origin") -> str:
        """
        Retrieve the URL of a Git remote repository with enhanced error handling.

        Args:
            remote_name (str): Name of the remote to query (default: 'origin')

        Returns:
            str: Remote URL if successful, 'NA' if failed to retrieve

        Features:
            - Detailed error categorization
            - Cross-platform compatibility
            - Safe subprocess execution
        """
        try:
            # Execute git command with timeout protection
            result = subprocess.run(
                ["git", "remote", "get-url", "origin"],
                check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                universal_newlines=True,
                encoding='utf-8',
                timeout=60  # 5-second timeout
            )

            url = result.stdout.strip() if result.stdout else None
            if not url:
                raise ValueError("Empty response from git command")

            Util.reportNormal(
                f"[Git Remote] Retrieved URL: {url}"
            )
            return url

        except subprocess.CalledProcessError as e:
            error_msg = (
                f"[Git Remote] Command failed ({e.returncode}): "
                f"{e.stderr.strip() or 'No error message'}"
            )
            Util.reportErrorNoExit(error_msg)

        except FileNotFoundError:
            Util.reportErrorNoExit(
                "[Git Remote] Git executable not found"
            )

        except subprocess.TimeoutExpired:
            Util.reportErrorNoExit(
                "[Git Remote] Operation timed out"
            )

        except PermissionError:
            Util.reportErrorNoExit(
                "[Git Remote] Permission denied"
            )

        except Exception as e:
            Util.reportErrorNoExit(
                f"[Git Remote] Unexpected error: {type(e).__name__} - {str(e)}"
            )

        return "NA"

    def get_git_current_branch(self) -> str:
        """Retrieve current Git branch name with enhanced diagnostics and fallback methods"""
        def report_success(branch):
            Util.reportNormal(f"[Git Remote] branch: {branch}")
            return branch

        # 方法1：使用 git branch --show-current
        try:
            result = subprocess.run(
                ["git", "branch", "--show-current"],
                check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                encoding='utf-8',
                timeout=30
            )
            if (branch := result.stdout.strip()) if result.stdout else None:
                return report_success(branch)
        except (subprocess.CalledProcessError, FileNotFoundError) as e:
            error_type = "Command failed" if isinstance(e, subprocess.CalledProcessError) else "Git not installed"
            error_msg = getattr(e, 'stderr', str(e))
            # 如果是字符串就直接使用，不需要decode
            if isinstance(error_msg, str):
                error_msg = error_msg[:200]
            else:
                error_msg = error_msg.decode()[:200]
            Util.reportErrorNoExit(
                f"[Git Branch] Method1 ({error_type}): {error_msg}"
            )

        # 方法2：使用 git symbolic-ref --short HEAD
        try:
            result = subprocess.run(
                ["git", "symbolic-ref", "--short", "HEAD"],
                check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                encoding='utf-8',
                timeout=30
            )
            if (branch := result.stdout.strip()) if result.stdout else None:
                return report_success(branch)
        except subprocess.CalledProcessError as e:
            Util.reportErrorNoExit(
                f"[Git Branch] Method2 (Code {e.returncode}): {e.stderr.strip()[:200]}"
            )

        # 方法3：解析.git/HEAD文件
        try:
            git_head_path = ".git/HEAD"
            # 处理子模块情况
            if not os.path.exists(git_head_path) and os.path.isfile(".git"):
                with open(".git", "r") as f:
                    content = f.read().strip()
                    if content.startswith("gitdir: "):
                        git_head_path = os.path.join(content[8:].strip(), "HEAD")

            if os.path.exists(git_head_path):
                with open(git_head_path, "r") as f:
                    content = f.read().strip()
                    if content.startswith("ref: refs/heads/"):
                        if branch := content.split("/")[-1]:
                            return report_success(branch)
        except Exception as e:
            Util.reportErrorNoExit(
                f"[Git Branch] Method3 Error: {str(e)[:200]}\nTrace: {traceback.format_exc()[-500:]}"
            )

        Util.reportErrorNoExit("[Git Branch] All methods failed to detect branch")
        return "NA"

    def filter_sup_by_file(self, suppression):
        return suppression.type != sType.INTERACTIVE_SUPPRESSION and suppression.isContinuous is False

    def sort_line_start(self, suppression):
        return suppression.lineStart

    def sort_count_message(self, m):
        return '%07d%s%04d' % (int(m.count), m.component, m.number)

    # Filter out suppressions we are not interested in.
    # Combine suppressions defined at line zero (since these are no longer combined in QAF).
    def filter_combine_suppressions(self, suppressions, suppression_filter):
        result = []
        for s in suppressions:
            if suppression_filter(s) is False:
                continue
            if len(s.suppressed) == 0:
                continue
            if s.lineStart != 0:
                # Add suppression to list as is.
                result.append(s)
                continue
            # Merge line zero suppressions.
            found = False
            for existing in result:
                if existing.location == s.location:
                    existing.suppressed.extend(s.suppressed)
                    existing.messages += ', ' + s.messages
                    found = True
                    break
            if found is False:
                result.append(s)

        return result

    def findFileKeyStatus(self, suppressPath, fullAnalysisPath, projectPath):
        # Suppress File Check
        suffixesSourceLower = [s.lower() for s in suffixesSource]
        suffixesIncludeLower = [s.lower() for s in suffixesInclude]
        prefixesKeyLower = [s.lower() for s in prefixesKey]

        suppressEnable = False
        findKey = False
        for suppress in suppressPath:
            suppress = suppress.rstrip('\n')

            if suppress.lower() in projectPath.replace('\\', '/').rstrip('\n').lower():
                suppressEnable = True
                break

        if suppressEnable == False:
            fileName = os.path.basename(projectPath).lower()

            if 'memmap' not in fileName:
                if (analysisMode == "full"):
                    for fullAnalysis in fullAnalysisPath:
                        fullAnalysis = fullAnalysis.rstrip('\n')
                        if fullAnalysis.lower() in projectPath.replace('\\', '/').rstrip('\n').lower():
                            findKey = True
                            break

                else:
                    if '.c' in fileName:
                        for key in sourceKeyWords:
                            for suffix in suffixesSourceLower:
                                if (key + suffix) == fileName:
                                    findKey = True
                                else:
                                    # Filter prefix
                                    for prefix in prefixesKeyLower:
                                        if (prefix + key + suffix) == fileName:
                                            findKey = True
                                            break
                                if findKey == True:
                                    break
                            if findKey == True:
                                break

                    if '.h' in fileName:
                        for key in sourceKeyWords:
                            for suffix in suffixesIncludeLower:
                                if (key + suffix) == fileName:
                                    findKey = True
                                else:
                                    # Filter prefix
                                    for prefix in prefixesKeyLower:
                                        if (prefix + key + suffix) == fileName:
                                            findKey = True
                                            break
                                if findKey == True:
                                    break
                            if findKey == True:
                                break

        return findKey

    def getQACRules(self):
        wb = load_workbook(self.QACRules)

        # Get the '报备项' worksheet
        ws = wb['报备项']

        # Read the header row
        firstRow = next(ws.iter_rows(min_row=1, max_row=1))
        headers = [cell.value for cell in firstRow]

        # Read all rows of data(skip the header)
        ruleList = []
        ruleIndex = 0
        for index, header in enumerate(headers):
            if 'MISRA Rule No.' in header:
                ruleIndex = index

        for row in ws.iter_rows(min_row=2, values_only=True):
            rawRules = row[ruleIndex]

            for rawRule in re.split(r'[\\、]', str(rawRules)):
                rule = int(rawRule)
                if rule not in ruleList:
                    ruleList.append(rule)

        # Get the '模块白名单' worksheet
        ws = wb['模块白名单']

        # Read the header row
        firstRow = next(ws.iter_rows(min_row=1, max_row=1))
        headers = [cell.value for cell in firstRow]

        # Read all rows of data(skip the header)
        ruleModuleMap = {}
        ruleIndex = 0
        moduleIndex = 1
        for index, header in enumerate(headers):
            if 'MISRA Rule No.' in header:
                ruleIndex = index
            elif 'WhiteList Module' in header:
                moduleIndex = index

        for row in ws.iter_rows(min_row=2, values_only=True):
            rawRules = row[ruleIndex]
            rawModules = row[moduleIndex]

            rules = re.split(r'[\\、]', str(rawRules))
            modules = re.split(r'[\\、]', str(rawModules))

            for rawRule in rules:
                try:
                    rule = int(rawRule.strip())
                    if rule not in ruleModuleMap:
                        ruleModuleMap[rule] = set()
                    ruleModuleMap[rule].update(m.strip().lower() for m in modules if m.strip())
                except ValueError:
                    continue

        # Get 'CATL编码规则' worksheet
        ws = wb['CATL编码规则']

        # Read the header row
        firstRow = next(ws.iter_rows(min_row=1, max_row=1))
        headers = [cell.value for cell in firstRow]

        # Read all rows of data(skip the header)
        ruleCode = []
        ruleIndex = 0
        for index, header in enumerate(headers):
            if 'MISRA Rule No.' in header:
                ruleIndex = index
        for row in ws.iter_rows(min_row=2, values_only=True):
            rawRules = row[ruleIndex]

            for rawRule in re.split(r'[\\、]', str(rawRules)):
                rule = int(rawRule.strip())
                if rule not in ruleCode:
                    ruleCode.append(rule)

        # close excel file
        wb.close()
        return (ruleList, ruleModuleMap, ruleCode)

    def build_hmr_content(self):
        """Build HMR (Helix Metrics Report) content"""

        # File level metrics definition
        FILE_METRICS_RULES = {
            'STCDN': '>0.2',  # Comment to Code Ratio (CATL: >0.2)
            'COMF': '>0.2',  # Comment Density (HIS: >0.2)
            'VOCF': '[1,4]',  # Language Set (HIS: 0..4)
            'STNRA': '0'  # Should be 0
        }

        FILE_METRICS_ORDER = [
            'STCDN', 'COMF', 'STNRA', 'STBME', 'STDIF', 'STECT', 'STFCO', 'STFNC',
            'STM22', 'STM28', 'STM33', 'STSCT', 'STSHN', 'STTLN',
            'STTPP', 'STVAR', 'VOCF'
        ]

        FILE_METRICS_NAMES = {
            'STCDN': 'STCDN',
            'COMF': '*COMF',
            'STNRA': 'STNRA',
            'VOCF': '*VOCF',
            'STBME': 'STBME',
            'STDIF': 'STDIF',
            'STECT': 'STECT',
            'STFCO': 'STFCO',
            'STFNC': 'STFNC',
            'STM22': 'STM22',
            'STM28': 'STM28',
            'STM33': 'STM33',
            'STSCT': 'STSCT',
            'STSHN': 'STSHN',
            'STTLN': 'STTLN',
            'STTPP': 'STTPP',
            'STVAR': 'STVAR'
        }

        # Function level metrics definition
        FUNCTION_METRICS_RULES = {
            'STPTH': '[1,80]',  # Estimated Static Program Paths (HIS: 1..80)
            'STCYC': '[1,10]',  # Cyclomatic Complexity (HIS: 1..10)
            'STGTO': '0',  # Number of Goto statements (HIS: 0)
            'STCAL': '[0,7]',  # Number of Functions Called from Function (HSI: 0..7)
            'STPAR': '[0,5]',  # Number of Function Parameters (HIS: 0..5)
            'STMIF': '[0,4]',  # Deepest Level of Nesting (HIS: 0..4)
            'STM29': '[0,5]',  # Number of Functions Calling this Function (HIS: 0..5)
            'STST3': '[1,50]',  # Number of Statements in Function (variant 3) (HIS: 0..50)
            'STM19': '[0,1]',  # Number of Exit Points (HIS: 0..1)
            'STRET': '[0,1]'  # Number of Return Points in Function (HIS: 0..1)
        }

        FUNCTION_METRICS_ORDER = [
            'STPTH', 'STCYC', 'STGTO', 'STCAL', 'STPAR', 'STMIF', 'STM29',
            'STST3', 'STM19', 'STRET', 'STAV1', 'STKNT', 'STLCT', 'STLIN',
            'STM07', 'STSUB', 'STUNR', 'STUNV', 'STXLN'
        ]

        FUNCTION_METRICS_NAMES = {
            'STPTH': 'STPTH',
            'STCYC': 'STCYC',
            'STGTO': 'STGTO',
            'STCAL': 'STCAL',
            'STPAR': 'STPAR',
            'STMIF': 'STMIF',
            'STM29': 'STM29',
            'STST3': 'STST3',
            'STM19': 'STM19',
            'STRET': 'STRET',
            'STAV1': 'STAV1',
            'STKNT': 'STKNT',
            'STLCT': 'STLCT',
            'STLIN': 'STLIN',
            'STM07': 'STM07',
            'STSUB': 'STSUB',
            'STUNR': 'STUNR',
            'STUNV': 'STUNV',
            'STXLN': 'STXLN'
        }

        def parse_rule(rule_str):
            """Parse rule string"""
            if rule_str is None:
                return (None, None, 'none')

            if isinstance(rule_str, str):
                rule_str = rule_str.strip()
                if rule_str.startswith('>'):
                    try:
                        value = float(rule_str[1:])
                        return (value, None, 'greater_than')
                    except:
                        return (None, None, 'none')
                elif rule_str.startswith('[') and rule_str.endswith(']'):
                    try:
                        range_str = rule_str[1:-1]
                        if ',' in range_str:
                            min_val, max_val = map(float, range_str.split(','))
                            return (min_val, max_val, 'range')
                        else:
                            value = float(range_str)
                            return (value, value, 'equal')
                    except:
                        return (None, None, 'none')
                else:
                    try:
                        value = float(rule_str)
                        return (value, value, 'equal')
                    except:
                        return (None, None, 'none')
            else:
                try:
                    value = float(rule_str)
                    return (value, value, 'equal')
                except:
                    return (None, None, 'none')

        def check_pass(metric_key, value, rules_dict):
            """Check if metric passes"""
            rule_str = rules_dict.get(metric_key)
            if rule_str is None:
                return True  # Metrics without rules always pass

            min_val, max_val, check_type = parse_rule(rule_str)

            if check_type == 'none':
                return True

            if value in ['/', '', 'NA', None, 'Infinity', 'infinity', 'INFINITY']:
                return True

            try:
                float_value = float(value)
            except (ValueError, TypeError):
                # If not a number, check if it's Infinity
                if str(value).lower() == 'infinity':
                    return True
                return False

            if check_type == 'greater_than':
                return float_value > min_val
            elif check_type == 'range':
                return min_val <= float_value <= max_val
            elif check_type == 'equal':
                return float_value == min_val
            else:
                return True

        def parse_metric_value(value):
            """Parse metric value"""
            if value in ['/', '', None]:
                return '/'

            if isinstance(value, (int, float)):
                return float(value)
            elif isinstance(value, str):
                # Handle Infinity
                if value.lower() in ['infinity', 'inf']:
                    return 'Infinity'
                try:
                    # Try to extract numbers
                    match = re.search(r'[-+]?\d*\.?\d+', value)
                    if match:
                        return float(match.group())
                except:
                    pass
            return value

        def format_value_display(value):
            """Format value display"""
            if value == '/':
                return '/'

            if value == 'Infinity':
                return 'Infinity'

            try:
                float_value = float(value)
                if float_value.is_integer():
                    return str(int(float_value))
                else:
                    # Display up to 3 decimal places
                    formatted = f'{float_value:.3f}'
                    # Remove trailing zeros
                    if formatted.endswith('.000'):
                        return formatted[:-4]
                    elif formatted.endswith('00'):
                        return formatted[:-2]
                    elif formatted.endswith('0'):
                        return formatted[:-1]
                    return formatted
            except:
                return str(value)

        def get_cell_color(value, metric_key, rules_dict):
            """Get cell color - use lime and yellow according to template"""
            if value == '/' or value == 'Infinity':
                return ''

            if metric_key not in rules_dict:
                return ''

            is_ok = check_pass(metric_key, value, rules_dict)

            # Special handling: STGTO=0 is compliant
            if metric_key == 'STGTO':
                try:
                    if float(value) == 0:
                        is_ok = True
                except:
                    pass

            if is_ok:
                return 'bgcolor="lime"'
            else:
                return 'bgcolor="yellow"'

        # Main logic starts
        resultXmlData = Util.getBuildPath() + r'/CodeVerify/CodeStaticCheck/prqa/configs/Initial/reports/results_data.xml'

        try:
            # Check if file exists
            if not os.path.exists(resultXmlData):
                print(f"HSI Metrics data file not found: {resultXmlData}")
                return '<div class="sec"><h1><a name="HSIMetrics">5 HSI Metrics Report</a></h1></div><p>HSI Metrics data file not found.</p>'

            # Parse XML file
            try:
                tree = ET.parse(resultXmlData)
                root = tree.getroot()
            except ET.ParseError as e:
                print(f"XML parsing error: {e}")
                return '<div class="sec"><h1><a name="HSIMetrics">5 HSI Metrics Report</a></h1></div><p>XML parsing error.</p>'

            # Collect all data
            all_files_data = []
            file_functions_map = {}  # filename -> function list
            xml_dir = os.path.dirname(resultXmlData)

            # Find all dataroot nodes
            for dataroot in root.findall('dataroot'):
                if dataroot.get("type") != "per-file":
                    continue

                for file_elem in dataroot.findall('File'):
                    json_elem = file_elem.find('Json')
                    if json_elem is None:
                        continue

                    json_file_name = json_elem.text
                    if not json_file_name:
                        continue

                    # Build complete JSON file path
                    json_path = os.path.join(xml_dir, json_file_name)

                    if not os.path.exists(json_path):
                        print(f"Warning: JSON file does not exist: {json_path}")
                        continue

                    file_path = file_elem.get('path', '')
                    file_name = file_elem.get('name') or os.path.basename(file_path)

                    # Only process .c files
                    if not file_name.endswith('.c'):
                        continue

                    # Read JSON data
                    try:
                        with open(json_path, 'r', encoding='utf-8') as f:
                            json_data = json.load(f)
                    except Exception as e:
                        print(f"Warning: Unable to parse JSON file {json_path}: {e}")
                        continue

                    # Extract file level metrics
                    metrics = json_data.get('metrics', {})

                    # Get all file level metrics
                    file_metrics = {}
                    for metric_key in FILE_METRICS_ORDER:
                        if metric_key in ['STCDN', 'COMF', 'VOCF', 'STNRA']:
                            # These are metrics that need checking
                            file_metrics[metric_key] = parse_metric_value(metrics.get(metric_key, 0))
                        elif metric_key in ['STBME', 'STDIF', 'STECT', 'STFCO', 'STFNC',
                                            'STM22', 'STM28', 'STM33', 'STSCT', 'STSHN',
                                            'STTLN', 'STTPP', 'STVAR']:
                            # These are display-only metrics
                            file_metrics[metric_key] = parse_metric_value(metrics.get(metric_key, 0))
                        else:
                            file_metrics[metric_key] = '/'

                    # Add to file data
                    all_files_data.append({
                        'name': file_name,
                        'metrics': file_metrics
                    })

                    # Extract function data
                    entities = json_data.get('entities', [])
                    file_functions = []
                    for entity in entities:
                        if entity.get('type') == 'function':
                            raw_func_name = entity.get('name', 'Unknown')
                            # Clean function name, remove parameter part
                            left_paren = raw_func_name.find('(')
                            if left_paren > 0:
                                func_name = raw_func_name[:left_paren].strip().lstrip('*')
                            else:
                                func_name = raw_func_name.strip()

                            func_metrics_data = entity.get('metrics', {})

                            func_metrics = {}
                            for metric_key in FUNCTION_METRICS_ORDER:
                                # All metrics that need checking
                                if metric_key in FUNCTION_METRICS_RULES:
                                    func_metrics[metric_key] = parse_metric_value(func_metrics_data.get(metric_key, 0))
                                # Display-only metrics
                                elif metric_key in ['STAV1', 'STKNT', 'STLCT', 'STLIN', 'STM07', 'STSUB', 'STUNR',
                                                    'STUNV', 'STXLN']:
                                    func_metrics[metric_key] = parse_metric_value(func_metrics_data.get(metric_key, 0))
                                else:
                                    func_metrics[metric_key] = '/'

                            # Add to function data
                            file_functions.append({
                                'function_name': func_name,
                                'metrics': func_metrics
                            })

                    # Sort by function name for easier viewing
                    file_functions.sort(key=lambda x: x['function_name'])
                    file_functions_map[file_name] = file_functions

            # Build HTML content
            html_content = []

            # 5 HSI Metrics Report title
            html_content.append('<div class="sec"><h1><a name="HSIMetrics">5 HSI Metrics Report</a></h1></div>')

            if not all_files_data:
                html_content.append('<p>No metrics data available for the analyzed files.</p>')
                return '\n'.join(html_content)

            # ============ File Metrics table ============
            html_content.append('<a name="File_Metrics"></a>')
            html_content.append('<table border="0" width="100%">')
            # File Metrics: 1 file column + 17 metric columns = 18 columns
            html_content.append('<colgroup>')
            html_content.append('<col width="20%" />')  # file column
            html_content.append('<col width="4.7%" />')  # STCDN
            html_content.append('<col width="4.7%" />')  # *COMF
            html_content.append('<col width="4.7%" />')  # STNRA
            html_content.append('<col width="4.7%" />')  # STBME
            html_content.append('<col width="4.7%" />')  # STDIF
            html_content.append('<col width="4.7%" />')  # STECT
            html_content.append('<col width="4.7%" />')  # STFCO
            html_content.append('<col width="4.7%" />')  # STFNC
            html_content.append('<col width="4.7%" />')  # STM22
            html_content.append('<col width="4.7%" />')  # STM28
            html_content.append('<col width="4.7%" />')  # STM33
            html_content.append('<col width="4.7%" />')  # STSCT
            html_content.append('<col width="4.7%" />')  # STSHN
            html_content.append('<col width="4.7%" />')  # STTLN
            html_content.append('<col width="4.7%" />')  # STTPP
            html_content.append('<col width="4.7%" />')  # STVAR
            html_content.append('<col width="4.7%" />')  # *VOCF
            html_content.append('</colgroup>')

            html_content.append('<tbody>')
            html_content.append('<tr bgcolor="#0028aa">')
            html_content.append('<td colspan="18"><font color="#ffffff" size="4">File Metrics</font> </td>')
            html_content.append('</tr>')
            html_content.append('<tr bgcolor="darkgray">')
            html_content.append('<td></td>')
            html_content.append('<th colspan="3">info-only metrics</th>')
            html_content.append('<th colspan="14"> measured-only metrics</th>')
            html_content.append('</tr>')
            html_content.append('<tr bgcolor="lightgray">')
            html_content.append('<td>file</td>')
            for metric_key in FILE_METRICS_ORDER:
                metric_name = FILE_METRICS_NAMES.get(metric_key, metric_key)
                html_content.append(f'<td>{metric_name}</td>')
            html_content.append('</tr>')

            # Display each file
            for file_data in all_files_data:
                file_name = file_data['name']
                html_content.append('<tr>')
                html_content.append(f'<td>{file_name}</td>')

                for metric_key in FILE_METRICS_ORDER:
                    value = file_data['metrics'].get(metric_key, '/')
                    display_value = format_value_display(value)
                    color_attr = get_cell_color(value, metric_key, FILE_METRICS_RULES)

                    if color_attr:
                        html_content.append(f'<td {color_attr}>{display_value}</td>')
                    else:
                        html_content.append(f'<td>{display_value}</td>')

                html_content.append('</tr>')

            html_content.append('</tbody>')
            html_content.append('</table>')

            # File metrics description table
            html_content.append('<table border="0" width="100%">')
            html_content.append('<colgroup><col width="1*" /><col width="1*" /><col width="1*" /></colgroup>')
            html_content.append('<tbody>')
            html_content.append('<tr>')
            html_content.append('<td border="0">')
            html_content.append('<small>')
            html_content.append('STBME: Embedded Programmer Months<br />')
            html_content.append('STCDN: Comment to Code Ratio (CATL: &gt;0.2)<br />')
            html_content.append('STDIF: Program Difficulty<br />')
            html_content.append('STECT: Number of External Variables Declared<br />')
            html_content.append('STFCO: Estimated Function Coupling<br />')
            html_content.append('STFNC: Number of Functions in File')
            html_content.append('</small>')
            html_content.append('</td>')
            html_content.append('<td border="0">')
            html_content.append('<small>')
            html_content.append('STNRA: NumBer of Recursions across the Project(HIS: =0)<br />')
            html_content.append('STM22: Number of Statements<br />')
            html_content.append('STM28: Number of Non-Header Comments<br />')
            html_content.append('STM33: Number of Internal Comments<br />')
            html_content.append('STSCT: Number of Static Variables Declared<br />')
            html_content.append('STSHN: Shannon Information Content<br />')
            html_content.append('STTLN: Total Preprocessed Source Lines')
            html_content.append('</small>')
            html_content.append('</td>')
            html_content.append('<td border="0">')
            html_content.append('<small>')
            html_content.append('STTPP: Total Unpreprocessed Code Lines<br />')
            html_content.append('STVAR: Total Number of Variables<br />')
            html_content.append('COMF: Comment Density (STM28/STM22) (HIS: &gt;0.2)<br />')
            html_content.append('VOCF: Language Set ((STM21+STM20)/(STOPT+STOPN)) (HIS: 0..4)')
            html_content.append('</small>')
            html_content.append('</td>')
            html_content.append('</tr>')
            html_content.append('</tbody>')
            html_content.append('</table>')
            html_content.append('<h1></h1>')

            # ============ Function Metrics table ============
            html_content.append('<a name="Function_Metrics"></a>')
            html_content.append('<table border="0" width="100%">')
            # Function Metrics: 1 function column + 19 metric columns = 20 columns
            # Adjust column width: function column takes 25%, other 19 columns take about 3.95% each
            html_content.append('<colgroup>')
            html_content.append('<col width="25%" />')  # function column
            html_content.append('<col width="3.95%" />')  # STPTH
            html_content.append('<col width="3.95%" />')  # STCYC
            html_content.append('<col width="3.95%" />')  # STGTO
            html_content.append('<col width="3.95%" />')  # STCAL
            html_content.append('<col width="3.95%" />')  # STPAR
            html_content.append('<col width="3.95%" />')  # STMIF
            html_content.append('<col width="3.95%" />')  # STM29
            html_content.append('<col width="3.95%" />')  # STST3
            html_content.append('<col width="3.95%" />')  # STM19
            html_content.append('<col width="3.95%" />')  # STRET
            html_content.append('<col width="3.95%" />')  # STAV1
            html_content.append('<col width="3.95%" />')  # STKNT
            html_content.append('<col width="3.95%" />')  # STLCT
            html_content.append('<col width="3.95%" />')  # STLIN
            html_content.append('<col width="3.95%" />')  # STM07
            html_content.append('<col width="3.95%" />')  # STSUB
            html_content.append('<col width="3.95%" />')  # STUNR
            html_content.append('<col width="3.95%" />')  # STUNV
            html_content.append('<col width="3.95%" />')  # STXLN
            html_content.append('</colgroup>')

            html_content.append('<tbody>')
            html_content.append('<tr bgcolor="#0028aa">')
            html_content.append('<td colspan="20"><font color="#ffffff" size="4">Function Metrics</font></td>')
            html_content.append('</tr>')
            html_content.append('<tr bgcolor="darkgray">')
            html_content.append('<th></th>')
            html_content.append('<th colspan="6">supervised metrics</th>')
            html_content.append('<th colspan="4">info-only metrics</th>')
            html_content.append('<th colspan="9"> measured-only metrics </th>')
            html_content.append('</tr>')
            html_content.append('<tr bgcolor="lightgray">')
            html_content.append('<td>function</td>')
            for metric_key in FUNCTION_METRICS_ORDER:
                metric_name = FUNCTION_METRICS_NAMES.get(metric_key, metric_key)
                html_content.append(f'<td>{metric_name}</td>')
            html_content.append('</tr>')

            # Collect all functions (sorted by function name)
            all_functions = []
            for file_data in all_files_data:
                file_name = file_data['name']
                functions = file_functions_map.get(file_name, [])

                for func_data in functions:
                    func_data_with_file = func_data.copy()
                    func_data_with_file['source_file'] = file_name
                    all_functions.append(func_data_with_file)

            # Sort by function name
            all_functions.sort(key=lambda x: x['function_name'])

            # Display all functions
            for func_data in all_functions:
                func_name = func_data['function_name']
                html_content.append('<tr>')
                html_content.append(f'<td>{func_name}</td>')

                for metric_key in FUNCTION_METRICS_ORDER:
                    value = func_data['metrics'].get(metric_key, '/')
                    display_value = format_value_display(value)
                    color_attr = get_cell_color(value, metric_key, FUNCTION_METRICS_RULES)

                    if color_attr:
                        html_content.append(f'<td {color_attr}>{display_value}</td>')
                    else:
                        html_content.append(f'<td>{display_value}</td>')

                html_content.append('</tr>')

            html_content.append('</tbody>')
            html_content.append('</table>')

            # Function metrics description table
            html_content.append('<table border="0" width="100%">')
            html_content.append('<colgroup><col width="1*" /><col width="1*" /><col width="1*" /></colgroup>')
            html_content.append('<tbody>')
            html_content.append('<tr>')
            html_content.append('<td border="0">')
            html_content.append('<small>')
            html_content.append('STAV1: Average Size of Statement in Function (variant 1)<br />')
            html_content.append('STCAL: Number of Functions Called from Function (HSI: 0..7)<br />')
            html_content.append('STCYC: Cyclomatic Complexity (HIS: 1..10)<br />')
            html_content.append('STGTO: Number of Goto statements (HIS: 0)<br />')
            html_content.append('STKNT: Knot Count<br />')
            html_content.append('STLCT: Number of Local Variables Declared<br />')
            html_content.append('STLIN: Number of Code Lines')
            html_content.append('</small>')
            html_content.append('</td>')
            html_content.append('<td border="0">')
            html_content.append('<small>')
            html_content.append('STM07: Essential Cyclomatic Complexity<br />')
            html_content.append('STM19: Number of Exit Points (HIS: 0..1)<br />')
            html_content.append('STM29 Number of Functions Calling this Function (HIS: 0..5)<br />')
            html_content.append('STMIF: Deepest Level of Nesting(HIS: 0..4)<br />')
            html_content.append('STPAR: Number of Function Parameters (HIS: 0..5)<br />')
            html_content.append('STPTH: Estimated Static Program Paths (HIS: 1..80)<br />')
            html_content.append('STRET: Number of Return Points in Function (HIS: 0..1)')
            html_content.append('</small>')
            html_content.append('</td>')
            html_content.append('<td border="0">')
            html_content.append('<small>')
            html_content.append('STST3: Number of Statements in Function (variant 3) (HIS: 0..50)<br />')
            html_content.append('STSUB: Number of Function Calls<br />')
            html_content.append('STUNR: Number of Unreachable Statements<br />')
            html_content.append('STUNV: Unused or Non-Reused Variables<br />')
            html_content.append('STXLN: Number of Executable Lines')
            html_content.append('</small>')
            html_content.append('</td>')
            html_content.append('</tr>')
            html_content.append('</tbody>')
            html_content.append('</table>')
            html_content.append('<h1></h1>')

            return '\n'.join(html_content)

        except PermissionError as e:
            print(f"Permission error: {e}")
            return '<div class="sec"><h1><a name="HSIMetrics">5 HSI Metrics Report</a></h1></div><p>Permission error: Unable to access required files.</p>'
        except OSError as e:
            print(f"Operating system error: {e}")
            return '<div class="sec"><h1><a name="HSIMetrics">5 HSI Metrics Report</a></h1></div><p>Operating system error.</p>'
        except Exception as e:
            print(f"Error processing HSI Metrics data: {str(e)}")
            import traceback
            traceback.print_exc()
            return '<div class="sec"><h1><a name="HSIMetrics">5 HSI Metrics Report</a></h1></div><p>Error processing HSI Metrics data.</p>'

    def extract_and_build_justifications(self):
        """Extract suppression information and build Justifications section"""

        def extract_suppressions_from_file(file_path):
            """Extract suppression information from single .c/.h file - line by line version"""
            suppressions = []

            try:
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    lines = f.readlines()

                i = 0
                while i < len(lines):
                    line = lines[i].rstrip('\n')

                    # Find suppression marker (***_**_** format - at least two underscores)
                    md_match = re.match(r'^\s*([A-Za-z0-9]+_[A-Za-z0-9]+_[A-Za-z0-9_]+):\s*$', line)
                    if md_match:
                        name = md_match.group(1)
                        reason_lines = []
                        risk_lines = []
                        prevention_lines = []
                        current_section = None
                        found_period_newline = False

                        i += 1
                        while i < len(lines):
                            current_line = lines[i].rstrip('\n').strip()

                            if re.match(r'^\s*[A-Za-z0-9]+_[A-Za-z0-9]+_[A-Za-z0-9_]+:', current_line):
                                break

                            if current_line.startswith('Reason:'):
                                current_section = 'reason'
                                reason_text = current_line[7:].strip()
                                if reason_text:
                                    reason_lines.append(reason_text)
                            elif current_line.startswith('Risk:'):
                                current_section = 'risk'
                                risk_text = current_line[5:].strip()
                                if risk_text:
                                    risk_lines.append(risk_text)
                            elif current_line.startswith('Prevention:'):
                                current_section = 'prevention'
                                prevention_text = current_line[11:].strip()
                                if prevention_text:
                                    prevention_lines.append(prevention_text)
                            elif current_section and current_line:
                                if current_section == 'reason':
                                    reason_lines.append(current_line)
                                elif current_section == 'risk':
                                    risk_lines.append(current_line)
                                elif current_section == 'prevention':
                                    prevention_lines.append(current_line)

                            if current_section == 'prevention' and prevention_lines:
                                last_line = prevention_lines[-1]
                                if last_line.endswith('.'):
                                    check_next = i + 1
                                    while check_next < len(lines):
                                        next_line_content = lines[check_next].strip()
                                        if not next_line_content:
                                            i = check_next
                                            found_period_newline = True
                                            break
                                        elif next_line_content.startswith('/*') or next_line_content.startswith('*/'):
                                            i = check_next - 1
                                            found_period_newline = True
                                            break
                                        elif re.match(r'^\s*[A-Za-z0-9]+_[A-Za-z0-9]+_[A-Za-z0-9_]+:',
                                                      next_line_content):
                                            i = check_next - 1
                                            found_period_newline = True
                                            break
                                        elif next_line_content.startswith('Reason:') or next_line_content.startswith(
                                                'Risk:') or next_line_content.startswith('Prevention:'):
                                            i = check_next - 1
                                            found_period_newline = True
                                            break
                                        else:
                                            check_next += 1
                                            continue

                                    if found_period_newline:
                                        break

                            if found_period_newline:
                                break

                            i += 1

                        reason = ' '.join(reason_lines) if reason_lines else "Not specified"
                        risk = ' '.join(risk_lines) if risk_lines else "Not specified"
                        prevention = ' '.join(prevention_lines) if prevention_lines else "Not specified"

                        def clean_section(text):
                            if not text or text == "Not specified":
                                return "Not specified"

                            patterns_to_remove = [
                                r'/\*.*',
                                r'\*/.*',
                                r'//.*',
                            ]

                            for pattern in patterns_to_remove:
                                text = re.sub(pattern, '', text)

                            text = re.sub(r'\s*\*\s*', ' ', text)
                            text = re.sub(r'\s+', ' ', text)
                            text = text.strip()

                            return text

                        reason = clean_section(reason)
                        risk = clean_section(risk)
                        prevention = clean_section(prevention)

                        if re.search(r'[A-Za-z0-9]+_[A-Za-z0-9]+_[A-Za-z0-9_]+', prevention):
                            sup_match = re.search(r'([A-Za-z0-9]+_[A-Za-z0-9]+_[A-Za-z0-9_]+)', prevention)
                            if sup_match:
                                split_pos = prevention.find(sup_match.group(1))
                                if split_pos > 0:
                                    prevention = prevention[:split_pos].strip()
                                else:
                                    prevention = "Not specified"

                        suppressions.append({
                            'name': name,
                            'reason': reason,
                            'risk': risk,
                            'prevention': prevention,
                            'file': os.path.basename(file_path)
                        })
                    else:
                        i += 1

            except Exception as e:
                print(f"Error extracting suppression information from file {file_path}: {str(e)}")

            return suppressions

        def get_qac_files():
            """Get all files actually analyzed by QAC (.c and related .h files)"""
            qac_files = []

            # Read the actual QAC project XML to get files being analyzed
            prqa_file = Util.getTargetPRQA()
            if os.path.exists(prqa_file):
                try:
                    tree = ET.parse(prqa_file)
                    root = tree.getroot()

                    for file_elem in root.findall('.//file'):
                        file_name = file_elem.get('name', '')
                        folder = file_elem.get('folder', '')

                        if file_name and folder:
                            folder = folder.replace('${SOURCE_ROOT}', Util.getProjectPath())
                            c_file_path = os.path.join(folder, file_name)

                            if os.path.exists(c_file_path):
                                qac_files.append(c_file_path)

                                base_name = os.path.splitext(file_name)[0]
                                dir_path = os.path.dirname(c_file_path)

                                suffixesIncludeLower = [s.lower() for s in suffixesInclude]
                                for suffix in suffixesIncludeLower:
                                    possible_header = os.path.join(dir_path, base_name + suffix)
                                    if os.path.exists(possible_header):
                                        qac_files.append(possible_header)

                                for include_elem in root.findall('.//include'):
                                    include_path = include_elem.get('path', '')
                                    if include_path:
                                        if '${SOURCE_ROOT}' in include_path:
                                            include_path = include_path.replace('${SOURCE_ROOT}', Util.getProjectPath())

                                        if os.path.exists(include_path) and os.path.isdir(include_path):
                                            for suffix in suffixesIncludeLower:
                                                possible_header = os.path.join(include_path, base_name + suffix)
                                                if os.path.exists(possible_header):
                                                    qac_files.append(possible_header)

                except Exception as e:
                    print(f"Error reading PRQA project file: {str(e)}")

            if not qac_files:
                print("Warning: Could not read PRQA project file, using fallback method")

                suffixesSourceLower = [s.lower() for s in suffixesSource]
                suffixesIncludeLower = [s.lower() for s in suffixesInclude]

                suppressPath = []
                lineKey = False
                with open(Util.getQacCheckFilePath(), 'r', encoding='utf-8') as f:
                    for line in f.readlines():
                        line = line.replace("\\", "/")
                        if not lineKey:
                            if line.startswith('[SuppressPath]'):
                                lineKey = True
                                continue
                        elif lineKey and line.startswith(r'['):
                            break
                        else:
                            if not line.startswith('./'):
                                continue
                            suppressPath.append(line[2:].rstrip('\n'))

                lineKey = False
                fullAnalysisPath = []
                with open(Util.getQacCheckFilePath(), 'r', encoding='utf-8') as f:
                    lines = f.readlines()
                    for line in lines:
                        line = line.replace("\\", "/")
                        if not lineKey:
                            if not line.startswith(r"["):
                                continue
                            if line.startswith(r"[FullAnalysis]"):
                                lineKey = True
                        elif lineKey and line.startswith(r"["):
                            break
                        else:
                            if not line.startswith('./'):
                                continue
                            fullAnalysisPath.append(line[2:].rstrip('\n'))

                projectSource = []
                if (compilerEnv == 'cmake'):
                    with open(os.path.join(Util.getBuildPath(), 'Cmake/source.cmake'), 'r', encoding='utf-8') as f:
                        projectSource = [i.replace('\\', '/') for i in
                                         re.compile(r'set\(source(.*?)\)', re.S | re.M).findall(f.read())[0].split('\n')
                                         if
                                         i.replace(' ', '') != '']
                elif (compilerEnv == 'emake'):
                    with open(os.path.join(Util.getBuildPath(), 'cfg/file/build_src.txt'), 'r', encoding='utf-8') as f:
                        projectSource = [re.compile(r'^(.*?)(?:\|1)?$').sub(r'\1', line.strip()).replace('\\', '/') for
                                         line in f.readlines()]

                for sourceLine in projectSource:
                    sourceLine = sourceLine.replace("\\", "/")
                    if not sourceLine.startswith('./'):
                        continue

                    suppressEnable = False
                    for suppress in suppressPath:
                        if suppress.lower() in sourceLine.lower():
                            suppressEnable = True
                            break

                    if not suppressEnable:
                        in_analysis_scope = False

                        if analysisMode == "full":
                            for fullAnalysis in fullAnalysisPath:
                                if fullAnalysis.lower() in sourceLine.lower():
                                    in_analysis_scope = True
                                    break
                        elif analysisMode == "guard":
                            if sourceKeyWords:
                                file_name_lower = os.path.basename(sourceLine).lower()
                                for key in sourceKeyWords:
                                    key_lower = key.lower()
                                    for suffix in suffixesSourceLower:
                                        if (key_lower + suffix) == file_name_lower:
                                            in_analysis_scope = True
                                            break
                                        for prefix in ['schm_', 'rte_']:
                                            if (prefix + key_lower + suffix) == file_name_lower:
                                                in_analysis_scope = True
                                                break
                                    if in_analysis_scope:
                                        break

                        if in_analysis_scope:
                            fileOriginPath = sourceLine.replace('\n', '').replace('./', Util.getSourcePath(), 1)
                            filePath = self.correct_case_path(fileOriginPath)

                            if os.path.exists(filePath) and filePath.lower().endswith('.c'):
                                qac_files.append(filePath)

                                base_name = os.path.splitext(os.path.basename(filePath))[0]
                                dir_path = os.path.dirname(filePath)

                                for suffix in suffixesIncludeLower:
                                    possible_header = os.path.join(dir_path, base_name + suffix)
                                    if os.path.exists(possible_header):
                                        qac_files.append(possible_header)

            return list(set(qac_files))

        # Main logic
        all_suppressions = []
        qac_files = get_qac_files()

        for file_path in qac_files:
            file_suppressions = extract_suppressions_from_file(file_path)
            all_suppressions.extend(file_suppressions)

        # Build HTML content
        html_content = []

        html_content.append('<h1></h1>')
        html_content.append(
            '<div class="sec"><h1><a name="Justifications">6 Justifications for QAC Suppressions</a></h1></div>')

        html_content.append('<table border="0" width="100%">')
        html_content.append('<colgroup><col width="15%"><col width="30%"><col width="25%"><col width="30%"></colgroup>')
        html_content.append('<tbody>')

        html_content.append('<tr>')
        html_content.append('<td colspan="10" bgcolor="#0028aa">')
        html_content.append('<font color="#ffffff" size="4">Justifications</font>')
        html_content.append('</td>')
        html_content.append('</tr>')

        html_content.append('<tr bgcolor="darkgray">')
        html_content.append('<th>name</th>')
        html_content.append('<th>reason</th>')
        html_content.append('<th>potential risks</th>')
        html_content.append('<th>preventionof risks</th>')
        html_content.append('</tr>')

        if not all_suppressions:
            html_content.append('<tr>')
            html_content.append('<td colspan="4" align="center">No justifications found in analyzed source files.</td>')
            html_content.append('</tr>')
        else:
            all_suppressions.sort(key=lambda x: x['name'])

            for sup in all_suppressions:
                def clean_text(text, max_length=400):
                    if not text or text == "Not specified":
                        return "Not specified"
                    text = re.sub(r'\s+', ' ', text).strip()
                    if len(text) > max_length:
                        return text[:max_length] + "..."
                    return text

                name = sup['name']
                reason = clean_text(sup['reason'])
                risk = clean_text(sup['risk'])
                prevention = clean_text(sup['prevention'])

                html_content.append('<tr>')
                html_content.append(f'<td valign="top"><a name="{name}">{name}</a></td>')
                html_content.append(f'<td valign="top">{reason}</td>')
                html_content.append(f'<td valign="top">{risk}</td>')
                html_content.append(f'<td valign="top">{prevention}</td>')
                html_content.append('</tr>')

        html_content.append('</tbody>')
        html_content.append('</table>')
        html_content.append('<h1></h1>')

        return '\n'.join(html_content)

    def build_qac_rules_summary(self):
        """Build QAC rule violation summary table - one row per rule, files displayed separately"""
        resultXmlData = Util.getBuildPath() + r'/CodeVerify/CodeStaticCheck/prqa/configs/Initial/reports/results_data.xml'

        try:
            # Get rule information from Excel
            ruleList, ruleModuleMap, ruleCode = self.getQACRules()

            # Parse XML data
            if not os.path.exists(resultXmlData):
                print(f"Result XML file not found: {resultXmlData}")
                return '<table border="0" width="100%"><colgroup><col width="5%"><col width="5%"><col width="5%"><col width="35%"><col width="10%"><col width="5%"><col width="30%"><col width="5%"></colgroup><tbody><tr bgcolor="#0028aa"><td colspan="8"><font color="#ffffff" size="4">summary</font></td></tr><tr bgcolor="darkgray"><th>rules</th><th>occurrence</th><th>msgId</th><th>description</th><th>file</th><th>occurrence</th><th>justification</th><th>result</th></tr><tr><td colspan="8">Result XML file not found</td></tr></tbody></table>'

            # Read and parse XML file
            print(f"Reading XML file: {resultXmlData}")
            tree = ET.parse(resultXmlData)
            root = tree.getroot()

            # Determine file scope for analysis
            global sourceKeyWords

            # Function to check if file is in analysis scope
            def is_file_in_analysis_scope(file_path, source_keywords):
                if not source_keywords:
                    return True

                file_name = os.path.basename(file_path).lower()

                for key in source_keywords:
                    key_lower = key.lower()

                    # Check .c files
                    if file_name == f"{key_lower}.c":
                        return True

                    c_suffixes = ['_cfg.c', '_lcfg.c', '_cbk.c', '_pbcfg.c',
                                  '_define.c', '_callout.c', '_callout_stubs.c',
                                  '_hal_core.c', '_irq.c']
                    for suffix in c_suffixes:
                        if file_name == f"{key_lower}{suffix}":
                            return True

                    c_prefixes = ['schm_', 'rte_']
                    for prefix in c_prefixes:
                        if file_name.startswith(prefix) and file_name.endswith(f"{key_lower}.c"):
                            return True
                        for suffix in c_suffixes:
                            if file_name.startswith(prefix) and file_name.endswith(f"{key_lower}{suffix}"):
                                return True

                # Check .h files
                for key in source_keywords:
                    key_lower = key.lower()

                    if file_name == f"{key_lower}.h":
                        return True

                    h_suffixes = ['_cfg.h', '_lcfg.h', '_cbk.h', '_pbcfg.h',
                                  '_define.h', '_callout.h', '_callout_stubs.h',
                                  '_hal_core.h', '_generaltypes.h', '_type.h',
                                  '_types.h', '_memmap.h', '_int.h', '_irq.h']
                    for suffix in h_suffixes:
                        if file_name == f"{key_lower}{suffix}":
                            return True

                    h_prefixes = ['schm_', 'rte_']
                    for prefix in h_prefixes:
                        if file_name.startswith(prefix) and file_name.endswith(f"{key_lower}.h"):
                            return True
                        for suffix in h_suffixes:
                            if file_name.startswith(prefix) and file_name.endswith(f"{key_lower}{suffix}"):
                                return True

                return False

            # Data structures
            rule_violations = {}
            file_violations = {}

            # Parse per-file level information
            for dataroot in root.findall('dataroot'):
                if dataroot.get('type') == 'per-file':
                    for file_elem in dataroot.findall('File'):
                        file_path = file_elem.get('path', '')
                        file_name = os.path.basename(file_path)

                        if not is_file_in_analysis_scope(file_path, sourceKeyWords):
                            continue

                        # Parse file's levels tree
                        file_levels_tree = file_elem.find('.//tree[@type="levels"]')
                        if file_levels_tree is None:
                            continue

                        # Extract active messages (active > 0)
                        active_messages = {}
                        target_levels = ['QA_WARNING', 'QA_USERMESSAGE']

                        for level in file_levels_tree.findall('.//Level'):
                            level_guid = level.get('guid', '')

                            if level_guid not in target_levels:
                                continue

                            for component in level.findall('.//Component'):
                                component_guid = component.get('guid', '')
                                component_name = component.get('name', '')

                                for msg in component.findall('.//Message'):
                                    msg_guid = msg.get('guid', '')
                                    msg_text = msg.get('text', '')
                                    msg_active = int(msg.get('active', 0))

                                    if msg_active <= 0:
                                        continue

                                    # Extract msg_id
                                    msg_id = ""
                                    if '-' in msg_guid:
                                        parts = msg_guid.split('-')
                                        if len(parts) >= 3:
                                            msg_id = parts[-1]

                                    # Clean problem description text
                                    clean_text = msg_text
                                    if msg_id and clean_text.startswith(msg_id):
                                        clean_text = clean_text[len(msg_id):].lstrip(':').lstrip()
                                    clean_text = clean_text.lstrip('.')

                                    # Remove placeholders
                                    patterns_to_remove = [
                                        r'\s*\(\s*e\.g\.\s*\'%1s\'\s*\)',
                                        r'\s*e\.g\.\s*\'%1s\'',
                                        r'\s*\(\s*e\.g\.\s*\'%2s\'\s*\)',
                                        r'\s*\'%1s\'',
                                        r'\s*\'%2s\'',
                                        r'\s*\'%3s\'',
                                        r'\s*\'%s\'',
                                        r'\s*%1s',
                                        r'\s*%2s',
                                        r'\s*%s',
                                    ]

                                    for pattern in patterns_to_remove:
                                        clean_text = re.sub(pattern, '', clean_text)
                                    clean_text = re.sub(r'\s+', ' ', clean_text).strip()

                                    # Store active message
                                    active_messages[msg_guid] = {
                                        'msg_id': msg_id,
                                        'msg_text': clean_text,
                                        'active': msg_active,
                                        'component_guid': component_guid,
                                        'component_name': component_name,
                                        'full_msg_guid': msg_guid
                                    }

                        if not active_messages:
                            continue

                        # Build message to rule_id mapping
                        file_rules_tree = file_elem.find('.//tree[@type="rules"]')
                        if file_rules_tree is None:
                            continue

                        msg_to_rule = {}

                        for rule_group in file_rules_tree.findall('.//RuleGroup'):
                            for rule in rule_group.findall('.//Rule'):
                                rule_id = rule.get('id', '')

                                for msg in rule.findall('.//Message'):
                                    msg_guid = msg.get('guid', '')

                                    if msg_guid in active_messages:
                                        msg_to_rule[msg_guid] = rule_id

                        # Support all formats *_**_* (at least two underscores)
                        file_prqa_suppressions = {}  # key: rule_id_or_msg_id, value: list of (line_num, sup_name)
                        if os.path.exists(file_path):
                            try:
                                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                                    content = f.read()

                                # Support all formats *_**_*, at least two underscores separating three parts
                                pattern = r'/\*\s*PRQA\s+S\s+([A-Za-z0-9\-\.]+)\s*\*/\s*/\*\s*([A-Za-z0-9]+_[A-Za-z0-9]+_[A-Za-z0-9_]+)\s*\*/'

                                for match in re.finditer(pattern, content):
                                    prqa_target = match.group(
                                        1)  # Could be number (e.g., 6080) or rule ID (e.g., Dir-1.1)
                                    sup_name = match.group(2)  # e.g., MD_FaultM_Test3, CWED_FaultM_1277, XXX_YYY_ZZZ

                                    # Get line number
                                    lines_before = content[:match.start()].count('\n')
                                    line_num = lines_before + 1

                                    if prqa_target not in file_prqa_suppressions:
                                        file_prqa_suppressions[prqa_target] = []

                                    file_prqa_suppressions[prqa_target].append({
                                        'line': line_num,
                                        'sup_name': sup_name
                                    })

                            except Exception as e:
                                print(f"Error processing suppression markers in file {file_path}: {str(e)}")

                        # Process all active messages for current file
                        for msg_guid, msg_info in active_messages.items():
                            rule_id = msg_to_rule.get(msg_guid, 'Unknown')
                            msg_id = msg_info['msg_id']
                            clean_text = msg_info['msg_text']
                            file_active = msg_info['active']
                            full_msg_guid = msg_info['full_msg_guid']

                            rule_key = (rule_id, msg_guid)
                            file_key = (file_path, rule_id, msg_guid)

                            justification_text = '<span style="color: red;">未报备</span>'
                            has_suppression = False
                            justification_details = []

                            # Check Excel rules
                            try:
                                if msg_id and msg_id.replace('.', '', 1).isdigit():
                                    if '.' in msg_id:
                                        if msg_id in ruleList:
                                            justification_text = f'<span style="color: green;">已报备,报备项: {msg_id}</span>'
                                            has_suppression = True
                                    else:
                                        msg_num = int(msg_id)
                                        if msg_num in ruleList:
                                            justification_text = f'<span style="color: green;">已报备,报备项: {msg_num}</span>'
                                            has_suppression = True
                            except ValueError:
                                pass

                            if not has_suppression:
                                found_suppressions = []

                                # 1. Check if msg_id has PRQA suppression (e.g., 6080)
                                if msg_id and msg_id in file_prqa_suppressions:
                                    found_suppressions.extend(file_prqa_suppressions[msg_id])

                                # 2. Check if rule_id has PRQA suppression (e.g., Dir-1.1)
                                if rule_id and rule_id in file_prqa_suppressions:
                                    found_suppressions.extend(file_prqa_suppressions[rule_id])

                                # 3. If rule_id contains hyphen, check part after hyphen
                                if rule_id and '-' in rule_id:
                                    after_hyphen = rule_id.split('-')[-1]
                                    if after_hyphen in file_prqa_suppressions:
                                        found_suppressions.extend(file_prqa_suppressions[after_hyphen])

                                # Deduplicate: remove duplicates by line number and suppression name
                                unique_suppressions = []
                                seen = set()
                                for sup in found_suppressions:
                                    key = f"{sup['line']}_{sup['sup_name']}"
                                    if key not in seen:
                                        seen.add(key)
                                        unique_suppressions.append(sup)

                                if unique_suppressions:
                                    if len(unique_suppressions) == file_active:
                                        # Suppression count equals violation count, complete suppression
                                        for sup in unique_suppressions:
                                            sup_link = f'<a href="#{sup["sup_name"]}">{sup["sup_name"]}</a>'
                                            justification_details.append((sup['line'], sup_link))
                                        has_suppression = True
                                    elif len(unique_suppressions) > file_active:
                                        # More suppressions than violations, show first file_active
                                        for sup in unique_suppressions[:file_active]:
                                            sup_link = f'<a href="#{sup["sup_name"]}">{sup["sup_name"]}</a>'
                                            justification_details.append((sup['line'], sup_link))
                                        has_suppression = True
                                    else:
                                        # Fewer suppressions than violations, partial suppression
                                        for sup in unique_suppressions:
                                            sup_link = f'<a href="#{sup["sup_name"]}">{sup["sup_name"]}</a>'
                                            justification_details.append((sup['line'], sup_link))
                                        justification_text = f'<span style="color: orange;">抑制不足 ({len(unique_suppressions)}/{file_active})</span>'

                            # Store file information
                            if file_key not in file_violations:
                                file_violations[file_key] = {
                                    'file_path': file_path,
                                    'file_name': file_name,
                                    'rule_id': rule_id,
                                    'msg_guid': msg_guid,
                                    'msg_id': msg_id,
                                    'msg_text': clean_text,
                                    'active': file_active,
                                    'justification_text': justification_text,
                                    'has_suppression': has_suppression,
                                    'justification_details': justification_details,
                                    'full_msg_guid': full_msg_guid
                                }

                            # Store rule information
                            if rule_key not in rule_violations:
                                rule_violations[rule_key] = {
                                    'rule_id': rule_id,
                                    'msg_guid': msg_guid,
                                    'msg_id': msg_id,
                                    'msg_text': clean_text,
                                    'total_active': 0,
                                    'files': [],
                                    'full_msg_guid': full_msg_guid
                                }

                            rule_violations[rule_key]['total_active'] += file_active

                            file_found = False
                            for existing_file in rule_violations[rule_key]['files']:
                                if existing_file['file_name'] == file_name:
                                    existing_file['active'] += file_active
                                    file_found = True
                                    break

                            if not file_found:
                                file_info = {
                                    'file_name': file_name,
                                    'file_path': file_path,
                                    'active': file_active,
                                    'justification_text': justification_text,
                                    'has_suppression': has_suppression,
                                    'justification_details': justification_details,
                                    'full_msg_guid': full_msg_guid
                                }
                                rule_violations[rule_key]['files'].append(file_info)

            if not rule_violations:
                return '<table border="0" width="100%"><colgroup><col width="5%"><col width="5%"><col width="5%"><col width="35%"><col width="10%"><col width="5%"><col width="30%"><col width="5%"></colgroup><tbody><tr bgcolor="#0028aa"><td colspan="8"><font color="#ffffff" size="4">summary</font></td></tr><tr bgcolor="darkgray"><th>rules</th><th>occurrence</th><th>msgId</th><th>description</th><th>file</th><th>occurrence</th><th>justification</th><th>result</th></tr><tr><td colspan="8">No rule violations found.</td></tr></tbody></table>'

            # Generate HTML table
            html_content = []

            html_content.append('<h1></h1>')
            html_content.append('<table border="0" width="100%">')
            html_content.append('<colgroup>')
            html_content.append('<col width="5%">')  # rules
            html_content.append('<col width="5%">')  # occurrence
            html_content.append('<col width="5%">')  # msgId
            html_content.append('<col width="35%">')  # description
            html_content.append('<col width="10%">')  # file
            html_content.append('<col width="5%">')  # file occurrence
            html_content.append('<col width="30%">')  # justification
            html_content.append('<col width="5%">')  # result
            html_content.append('</colgroup>')
            html_content.append('<tbody>')
            html_content.append('<tr bgcolor="#0028aa">')
            html_content.append('<td colspan="8"><font color="#ffffff" size="4">summary</font></td>')
            html_content.append('</tr>')
            html_content.append('<tr bgcolor="darkgray">')
            html_content.append('<th>rules</th>')
            html_content.append('<th>occurrence</th>')
            html_content.append('<th>msgId</th>')
            html_content.append('<th>description</th>')
            html_content.append('<th>file</th>')
            html_content.append('<th>occurrence</th>')
            html_content.append('<th>justification</th>')
            html_content.append('<th>result</th>')
            html_content.append('</tr>')

            # Calculate overall result
            all_rules_passed = True
            for rule_key, rule_info in rule_violations.items():
                rule_passed = all(file.get('has_suppression', False)
                                  for file in rule_info.get('files', []))
                if not rule_passed:
                    all_rules_passed = False

            # Sort rules
            sorted_rules = sorted(rule_violations.items(),
                                  key=lambda x: (x[1]['rule_id'], x[1]['msg_id'] or ''))

            for rule_key, rule_info in sorted_rules:
                rule_id = rule_info['rule_id']
                msg_id = rule_info['msg_id']
                msg_text = rule_info['msg_text']
                total_active = rule_info['total_active']
                files = rule_info.get('files', [])
                full_msg_guid = rule_info.get('full_msg_guid', '')

                description = msg_text
                if len(description) > 100:
                    description = description[:97] + "..."

                msg_id_display = msg_id if msg_id else 'N/A'

                rule_passed = all(file.get('has_suppression', False) for file in files)
                result = 'ok' if rule_passed else 'failed'
                result_color = 'lime' if result == 'ok' else 'yellow'

                if len(files) == 1:
                    file_info = files[0]
                    html_content.append('<tr>')

                    html_content.append(f'<td>{rule_id}</td>')
                    html_content.append(f'<td>{total_active}</td>')
                    html_content.append(f'<td>{msg_id_display}</td>')
                    html_content.append(f'<td>{description}</td>')
                    html_content.append(f'<td>{file_info["file_name"]}</td>')
                    html_content.append(f'<td>{file_info["active"]}</td>')

                    justification_html = f'{file_info["file_name"]}'
                    if file_info['justification_details']:
                        justification_html += '<table border="0" width="100%">'
                        justification_html += '<colgroup><col width="10%"><col width="90%"></colgroup>'
                        justification_html += '<tbody>'
                        for line_num, sup_info in file_info['justification_details']:
                            justification_html += f'<tr><td>{line_num}</td><td>{sup_info}</td></tr>'
                        justification_html += '</tbody></table>'
                    elif file_info['justification_text'].find('已报备') != -1 or \
                            file_info['justification_text'].find('模块白名单') != -1 or \
                            file_info['justification_text'].find('CATL编码规则') != -1:
                        justification_html += f'<br>{file_info["justification_text"]}'
                    else:
                        justification_html += f'<br>{file_info["justification_text"]}'

                    html_content.append(f'<td>{justification_html}</td>')
                    html_content.append(f'<td bgcolor="{result_color}">{result}</td>')
                    html_content.append('</tr>')
                else:
                    first_file = True
                    for i, file_info in enumerate(files):
                        html_content.append('<tr>')
                        if first_file:
                            html_content.append(f'<td rowspan="{len(files)}">{rule_id}</td>')
                            html_content.append(f'<td rowspan="{len(files)}">{total_active}</td>')
                            html_content.append(f'<td rowspan="{len(files)}">{msg_id_display}</td>')
                            html_content.append(f'<td rowspan="{len(files)}">{description}</td>')
                            first_file = False

                        html_content.append(f'<td>{file_info["file_name"]}</td>')
                        html_content.append(f'<td>{file_info["active"]}</td>')

                        justification_html = f'{file_info["file_name"]}'
                        if file_info['justification_details']:
                            justification_html += '<table border="0" width="100%">'
                            justification_html += '<colgroup><col width="10%"><col width="90%"></colgroup>'
                            justification_html += '<tbody>'
                            for line_num, sup_info in file_info['justification_details']:
                                justification_html += f'<tr><td>{line_num}</td><td>{sup_info}</td></tr>'
                            justification_html += '</tbody></table>'
                        elif file_info['justification_text'].find('已报备') != -1 or \
                                file_info['justification_text'].find('模块白名单') != -1 or \
                                file_info['justification_text'].find('CATL编码规则') != -1:
                            justification_html += f'<br>{file_info["justification_text"]}'
                        else:
                            justification_html += f'<br>{file_info["justification_text"]}'

                        html_content.append(f'<td>{justification_html}</td>')

                        if i == 0:
                            html_content.append(f'<td rowspan="{len(files)}" bgcolor="{result_color}">{result}</td>')

                        html_content.append('</tr>')

            html_content.append('</tbody>')
            html_content.append('</table>')
            html_content.append('<h1></h1>')

            return '\n'.join(html_content)

        except Exception as e:
            import traceback
            error_msg = f"{str(e)}"
            traceback.print_exc()
            print(f"build_qac_rules_summary error: {error_msg}")

            return '<table border="0" width="100%"><colgroup><col width="5%"><col width="5%"><col width="5%"><col width="35%"><col width="10%"><col width="5%"><col width="30%"><col width="5%"></colgroup><tbody><tr bgcolor="#0028aa"><td colspan="8"><font color="#ffffff" size="4">summary</font></td></tr><tr bgcolor="darkgray"><th>rules</th><th>occurrence</th><th>msgId</th><th>description</th><th>file</th><th>occurrence</th><th>justification</th><th>result</th></tr><tr><td colspan="8">Processing error: ' + str(
                e)[:100] + '</td></tr></tbody></table>'

    def analysisReport(self):
        resultXmlData = Util.getBuildPath() + r'/CodeVerify/CodeStaticCheck/prqa/configs/Initial/reports/results_data.xml'
        resultsData = Util.getBuildPath() + r'/CodeVerify/CodeStaticCheck/prqa/configs/Initial/reports/results_data.pb2'
        outPutFile = Util.getBuildPath() + r'/CodeVerify/CodeStaticCheck/prqa/configs/Initial/reports/CATL_QAC_Test.html'
        ruleList, ruleModuleMap, ruleCode = self.getQACRules()

        suffixesSourceLower = [s.lower() for s in suffixesSource]
        suffixesIncludeLower = [s.lower() for s in suffixesInclude]
        prefixesKeyLower = [s.lower() for s in prefixesKey]

        # Suppress File Check
        suppressPath = []
        suppressData = []
        filesData = []
        sumSuppression = 0
        sumReportedSup = 0
        sumActive = 0

        # Read suppress analysis files
        lineKey = False
        with open(Util.getQacCheckFilePath(), 'r', encoding='utf-8')  as f:
            for i in f.readlines():
                i = i.replace("\\", "/")
                if not lineKey:
                    if i.startswith('[SuppressPath]'):
                        lineKey = True
                        continue
                elif lineKey == True and i.startswith(r'['):
                    break
                else:
                    if not i.startswith('./'):
                        continue
                    suppressPath.append(i[2:])

        # Read full analysis files
        lineKey = False
        fullAnalysisPath = []
        with open(Util.getQacCheckFilePath(), 'r', encoding='utf-8') as f:
            lines = f.readlines()
            for line in lines:
                line = line.replace("\\", "/")
                if not lineKey:
                    if not line.startswith(r"["):
                        continue
                    if line.startswith(r"[FullAnalysis]"):
                        lineKey = True
                elif lineKey and line.startswith(r"["):
                    break
                else:
                    if not line.startswith('./'):
                        continue
                    fullAnalysisPath.append(line[2:])

        tableContents = ''
        tableContents += '            <li><a href="#DiagnosticsSummary">4 Diagnostics Summary</a></li>\n'
        tableContents += '            <li><a href="#HSIMetrics">5 HSI Metrics Report</a></li>\n'
        tableContents += '            <li><a href="#Justifications">6 Justifications for QAC Suppressions</a></li>\n'

        # Collect suppression information from summary table
        summary_suppressions = {}  # file -> {rule -> suppression info}

        try:
            # Parse XML to get active message information
            tree = ET.parse(resultXmlData)
            root = tree.getroot()

            # Collect all file data
            allFilesData = []
            sumActive = 0
            sumReportedActive = 0

            # Find the dataroot node with type="per-file"
            dataRoot = root.find('.//dataroot[@type="per-file"]')

            # Fix FutureWarning: use explicit check
            if dataRoot is not None:
                for fileNode in dataRoot.findall('File'):
                    path = fileNode.get('path')
                    if not path:
                        continue

                    findKey = self.findFileKeyStatus(suppressPath, fullAnalysisPath, path)
                    if findKey == True:
                        activeCount = 0
                        reportActiveCount = 0

                        # Find all QA_USERMESSAGE and QA_WARNING Level nodes
                        for levelNode in fileNode.findall('.//Level'):
                            levelGuid = levelNode.get('guid')
                            if levelGuid in ['QA_USERMESSAGE', 'QA_WARNING']:

                                # Collect each Message
                                for messageNode in levelNode.findall('.//Message'):
                                    messageId = messageNode.get('guid')
                                    messageActive = messageNode.get('active')
                                    if messageId and messageActive and int(messageActive) > 0:
                                        number = messageId.split('-')[-1]
                                        isWhitelisted = False
                                        if int(number) in ruleList:
                                            isWhitelisted = True
                                            reportActiveCount += int(messageActive)

                                        if isWhitelisted != True:
                                            if int(number) in ruleModuleMap:
                                                for module in ruleModuleMap[int(number)]:
                                                    modulePattern = module.strip().replace('*', '.*')
                                                    if re.match(modulePattern, path.replace('\\', '/'), re.IGNORECASE):
                                                        isWhitelisted = True
                                                        reportActiveCount += int(messageActive)
                                                        break

                                        if isWhitelisted != True:
                                            if int(number) in ruleCode:
                                                isWhitelisted = True
                                                reportActiveCount += int(messageActive)

                                        activeCount += int(messageActive)

                        sumActive += activeCount
                        sumReportedActive += reportActiveCount

                        # Store file information
                        allFilesData.append({
                            "path": path,
                            "activeCount": activeCount,
                            "reportActiveCount": reportActiveCount,
                            "fileName": os.path.basename(path)
                        })

            # Collect suppression information using same logic as summary table
            # Re-parse each file to find PRQA suppression markers
            for fileInfo in allFilesData:
                filePath = fileInfo["path"]
                fileName = fileInfo["fileName"]

                if not os.path.exists(filePath):
                    continue

                # Read file content to find suppression markers
                try:
                    with open(filePath, 'r', encoding='utf-8', errors='ignore') as f:
                        content = f.read()

                    # Support all formats *_**_*, at least two underscores separating three parts
                    pattern = r'/\*\s*PRQA\s+S\s+([A-Za-z0-9\-\.]+)\s*\*/\s*/\*\s*([A-Za-z0-9]+_[A-Za-z0-9]+_[A-Za-z0-9_]+)\s*\*/'

                    file_suppressions = {}

                    for match in re.finditer(pattern, content):
                        prqa_target = match.group(1)  # Could be number (e.g., 6080) or rule ID (e.g., Dir-1.1)
                        sup_name = match.group(2)  # e.g., MD_FaultM_Test3, CWED_FaultM_1277, CD_xxx

                        # Get line number
                        lines_before = content[:match.start()].count('\n')
                        line_num = lines_before + 1

                        if prqa_target not in file_suppressions:
                            file_suppressions[prqa_target] = []

                        file_suppressions[prqa_target].append({
                            'line': line_num,
                            'sup_name': sup_name
                        })

                    summary_suppressions[fileName] = file_suppressions

                except Exception as e:
                    print(f"Error reading suppression markers from file {filePath}: {str(e)}")

            # Calculate suppression statistics for each file
            totalUnreportedSuppression = 0
            totalReportedSuppression = 0

            # Fix FutureWarning: need to check if dataRoot is None
            dataRoot = root.find('.//dataroot[@type="per-file"]')

            for fileInfo in allFilesData:
                fileName = fileInfo["fileName"]
                filePath = fileInfo["path"]
                activeCount = fileInfo["activeCount"]

                # Count suppression information for this file
                unreported_sup_count = 0
                reported_sup_count = 0

                # Get suppression information for this file
                file_suppressions = summary_suppressions.get(fileName, {})

                # Get active rule violations for this file from XML
                file_active_rules = {}
                if dataRoot is not None:
                    for fileNode in dataRoot.findall('File'):
                        if fileNode.get('path') == filePath:
                            for levelNode in fileNode.findall('.//Level'):
                                levelGuid = levelNode.get('guid')
                                if levelGuid in ['QA_USERMESSAGE', 'QA_WARNING']:
                                    for messageNode in levelNode.findall('.//Message'):
                                        messageId = messageNode.get('guid')
                                        messageActive = messageNode.get('active')
                                        if messageId and messageActive and int(messageActive) > 0:
                                            number = messageId.split('-')[-1]
                                            file_active_rules[number] = int(messageActive)

                # Match suppression markers with active rules
                for rule_id, sup_list in file_suppressions.items():
                    for sup_info in sup_list:
                        sup_name = sup_info['sup_name']

                        # Check if this suppression is effective (corresponds to an active rule)
                        is_effective = False

                        # 1. Suppression directly corresponds to rule ID
                        if rule_id in file_active_rules:
                            is_effective = True

                        # 2. Suppression corresponds to msg_id
                        elif rule_id.isdigit() and rule_id in file_active_rules:
                            is_effective = True

                        # 3. If rule_id contains hyphen, check part after hyphen
                        elif '-' in rule_id:
                            after_hyphen = rule_id.split('-')[-1]
                            if after_hyphen in file_active_rules:
                                is_effective = True

                        if is_effective:
                            # Check if suppression is already reported
                            is_reported = False

                            # Extract possible rule number from any ***_**_** format suppression
                            match = re.search(r'\d+', sup_name)
                            if match:
                                rule_num = int(match.group())
                                if rule_num in ruleList:
                                    is_reported = True
                                elif rule_num in ruleCode:
                                    is_reported = True
                                else:
                                    # Check module whitelist
                                    if rule_num in ruleModuleMap:
                                        for module in ruleModuleMap[rule_num]:
                                            modulePattern = module.strip().replace('*', '.*')
                                            if re.match(modulePattern, filePath.replace('\\', '/'), re.IGNORECASE):
                                                is_reported = True
                                                break

                            if is_reported:
                                reported_sup_count += 1
                            else:
                                unreported_sup_count += 1

                # Ensure suppression count does not exceed active issue count
                effective_unreported = min(unreported_sup_count, activeCount)
                effective_reported = min(reported_sup_count, activeCount)

                # Store file suppression information
                suppressData.append({
                    "path": filePath,
                    "fileName": fileName,
                    "unreportedSuppression": effective_unreported,
                    "reportedSuppression": effective_reported,
                    "activeCount": activeCount
                })

                totalUnreportedSuppression += effective_unreported
                totalReportedSuppression += effective_reported

            # Update total suppression statistics
            sumSuppression = totalUnreportedSuppression + totalReportedSuppression
            sumReportedSup = totalReportedSuppression

        except Exception as e:
            print(f"Error parsing suppression information: {str(e)}")
            import traceback
            traceback.print_exc()

        # Calculate remaining issues
        totalActive = sumActive - sumReportedActive  # Unreported QAC count
        totalSuppression = totalUnreportedSuppression  # Explained suppression count (unreported)
        totalReportedSup = totalReportedSuppression  # Reported suppression count

        # New judgment logic: Pass if total suppression >= unreported QAC count
        totalSuppressionAll = totalReportedSup + totalSuppression  # All suppression count
        insufficient_count = max(0, totalActive - totalSuppressionAll)  # Calculate insufficient count

        if insufficient_count == 0:
            resultText = f"Pass"
            resultColor = "green"
        else:
            resultText = f"Fail, Reported QAC count: {sumReportedActive}, Unreported QAC count: {totalActive}, Reported suppression count: {totalReportedSup}, Explained suppression count: {totalSuppression}"
            resultColor = "red"

        try:
            # Create XML report
            testsuites = ET.Element('testsuites')
            testsuite = ET.SubElement(testsuites, "testsuite")
            testsuite.set("name", "QAC verification result: " + resultText)

            # Total testcases
            testcaseNum = sumActive + totalSuppression
            if testcaseNum == 0:
                testcaseNum = 1
            testsuite.set("tests", str(testcaseNum))

            # Failed testcases
            testsuite.set("failures", str(totalSuppression))

            # Report testcases
            testsuite.set("errors", str(totalActive))

            # Skipped testcases
            totalSkip = sumReportedActive + totalReportedSup
            testsuite.set("skipped", str(totalSkip))

            filePath = Util.getBuildPath() + r'/CodeVerify/CodeStaticCheck/prqa/configs/Initial/reports/CATL_QAC_Report.xml'

            # Write XML
            tree = ET.ElementTree(testsuites)
            tree.write(filePath, encoding="utf-8", xml_declaration=True)

            # Generate HMR content
            hmr_content = self.build_hmr_content()

            # Generate Justifications section
            justifications_content = self.extract_and_build_justifications()

            # Generate Rule Summary section
            qac_rules_summary_content = self.build_qac_rules_summary()

            # Create HTML report
            with open(Util.getQacReportTemplatePath(), "r", encoding="utf-8") as f:
                htmlTemplate = f.read()

            # Build Chapter 4 "4 Diagnostics Summary" - includes Diagnostics Summary table and summary table
            diagSummary = []
            fileList = []
            row = '<h1></h1>\n'
            row += '  <div class="sec"><h1><a name="DiagnosticsSummary">4 Diagnostics Summary</a></h1></div>\n'
            row += '  <table border="0" width="100%">\n'
            row += '    <colgroup><col width="40%"><col width="15%"><col width="15%"><col width="15%"><col width="15%"></colgroup>\n'
            row += '    <tbody>\n'
            row += '      <tr bgcolor="#0028aa">\n'
            row += '        <td colspan="5"><font color="#ffffff" size="4">Diagnostics Summary</font> </td>\n'
            row += '      </tr>\n'
            diagSummary.append(row)
            row = f'<tr bgcolor="lightgray"><td>File</td><td>Active Diagnostics</td><td>Active Skip</td><td>Suppressed Diagnostics</td><td>Suppressed Skip</td></tr>'
            diagSummary.append(row)

            for fileInfo in allFilesData:
                filePath = fileInfo["path"]
                fileName = fileInfo["fileName"]
                fileList.append(fileName)
                activeCount = fileInfo["activeCount"]
                reportActiveCount = fileInfo["reportActiveCount"]

                # Find suppression information for this file
                unreported_sup = 0
                reported_sup = 0

                for suppressInfo in suppressData:
                    if fileName == suppressInfo["fileName"]:
                        unreported_sup = suppressInfo["unreportedSuppression"]
                        reported_sup = suppressInfo["reportedSuppression"]
                        break

                row = f'<tr><td align="left"><a href="{filePath}" title="{filePath}">{fileName}</a></td><td align="left">{activeCount}</td><td align="left">{reportActiveCount}</td><td align="left">{unreported_sup}</td><td align="left">{reported_sup}</td></tr>'
                diagSummary.append(row)

            row = f'<tr><td><strong>Total</strong></td><td align="left"><strong>{sumActive}</strong></td><td align="left"><strong>{sumReportedActive}</strong></td><td align="left"><strong>{totalUnreportedSuppression}</strong></td><td align="left"><strong>{totalReportedSuppression}</strong></td></tr>'
            diagSummary.append(row)
            row = f'<tr><td><strong>Result</strong></td><td colspan="4" align="left" style="color: {resultColor};"><strong>{resultText}</strong></td></tr>'
            diagSummary.append(row)

            row = f'</tbody>\n</table>\n\n'
            diagSummary.append(row)

            # Rule Summary display
            diagSummary.append(qac_rules_summary_content)

            # HSI Metrics display
            diagSummary.append(hmr_content)

            # Justifications display
            diagSummary.append(justifications_content)

            self.git_url = self.get_git_remote_url()
            self.git_branch = self.get_git_current_branch()

            if self.git_url == "NA":
                git_url = self.projectName
            else:
                git_url = self.git_url

            # Replace template placeholders
            htmlContent = htmlTemplate.format(tableContents=tableContents, testPerson=getpass.getuser(),
                                              testDate=datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
                                              fileList="<br>".join(fileList), projectPath=git_url,
                                              projectBranch=self.git_branch, buildPath=Util.getBuildPath(),
                                              cctFile=Util.getCCTPath() + '/' + self.generalConfig.get("qac", "cct"),
                                              diagSummary="\n".join(diagSummary))

            filePath = Util.getBuildPath() + r'/CodeVerify/CodeStaticCheck/prqa/configs/Initial/reports/CATL_QAC_Report.html'

            # Save HTML report
            with open(filePath, "w", encoding="utf-8") as f:
                f.write(htmlContent)

            print(f"Report saved to: {filePath}")

        except ET.ParseError as e:
            print(f"XML parsing error: {e}")
        except FileNotFoundError:
            print(f"XML file not found: {resultXmlData}")
        except Exception as e:
            print(f"Error occurred: {str(e)}, {type(e).__name__} at line {sys.exc_info()[-1].tb_lineno}")

if __name__ == '__main__':
    args = sys.argv[1:]
    modeParm = 'normal'
    analysisModeParm = 'specify'
    changeFileParm = ''
    for arg in args:
        if arg.startswith('-mode='):
            modeParm = arg.split('=')[1]
        elif arg.startswith('-analysismode='):
            analysisModeParm = arg.split('=')[1]
        elif arg.startswith('-file='):
            changeFileParm = arg.split('=')[1]
            if os.path.isfile(Util.getBuildPath() + changeFileParm):
                changeFile= Util.getBuildPath() + changeFileParm
            else:
                Util.reportError(f'-file={changeFileParm} not exist ')
                raise ValueError

    if modeParm == 'ci':
        workMode = 'ci'
    elif modeParm == 'normal':
        workMode = 'normal'
    else :
        Util.reportError(f'-mode={modeParm} not support ')
        raise ValueError

    if analysisModeParm == 'full':
        analysisMode = 'full'
    elif analysisModeParm == 'specify':
        analysisMode = 'specify'
    elif analysisModeParm == 'guard':
        analysisMode = 'guard'
    else:
        Util.reportError(f'-analysismode={analysisModeParm} not support ')
        raise ValueError

    Util.reportNormal(f'Current working mode: {workMode}')
    Util.reportNormal(f'Current analysis mode: {analysisMode}')
    Util.reportNormal(f'Current file: {changeFileParm}')

    # compiler select
    buildBatPath = Util.getBuildPath() + '/Tools/Build.bat'
    if(not os.path.exists(buildBatPath)):
        Util.reportError(buildBatPath + "->  build.bat don't exist")
    fd = open(buildBatPath, 'r', encoding='utf-8')
    text = fd.read()
    if(re.compile(r'Emake\.exe', re.IGNORECASE).findall(text) != []):
        compilerEnv = 'emake'
        Util.reportNormal(f"Found 'Emake.exe' in {buildBatPath}")

    Util.reportNormal(f"Compiler select: {compilerEnv}")

    qac = QacUtil()
    qac.start()